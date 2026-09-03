import csv
import json
import tempfile
import unittest
from pathlib import Path

from mxlperf.summary import build_summary, scan
from openpyxl import load_workbook


class SummaryTests(unittest.TestCase):
    def test_real_cpu_total_usage_is_averaged_and_rendered(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            result = root / "passing-run"
            result.mkdir()
            (result / "config.json").write_text(json.dumps({
                "SCENARIO": "pinned",
                "DESCRIPTION": "test",
                "PLACEMENT": "exclusive",
                "STREAMS": "1",
                "RESOLUTION": "1080p",
                "PRESET": "medium",
                "BITRATE": "12M",
                "ENC_CORES": "5",
                "ENC_THREADS": "15",
                "SLICES": "2",
            }))
            with (result / "metrics.csv").open("w", newline="") as handle:
                writer = csv.DictWriter(handle, fieldnames=["category", "metric", "value", "role"])
                writer.writeheader()
                writer.writerows([
                    {"category": "FFmpeg progress", "metric": "fps", "value": "avg=60.0;min=60.0;max=60.0", "role": "encoder"},
                    {"category": "CPU", "metric": "FFmpeg CPU demand", "value": "4.5", "role": "encoder"},
                    {"category": "CPU", "metric": "FFmpeg CPU demand", "value": "0.5", "role": "decoder"},
                    {"category": "CPU", "metric": "average FFmpeg utilization per measured used CPU", "value": "90.0", "role": "encoder"},
                    {"category": "Prometheus", "metric": "cross_numa_upi_incoming_bytes_per_second", "value": "avg=2500000000;min=2000000000;max=3000000000", "role": ""},
                ])
            with (result / "core-system-usage.csv").open("w", newline="") as handle:
                writer = csv.DictWriter(handle, fieldnames=["cpu_id", "real_cpu_total_usage_pct"])
                writer.writeheader()
                writer.writerows([
                    {"cpu_id": "0", "real_cpu_total_usage_pct": "10"},
                    {"cpu_id": "1", "real_cpu_total_usage_pct": "40"},
                    {"cpu_id": "2", "real_cpu_total_usage_pct": "90"},
                ])
            with (result / "cpu-usage.csv").open("w", newline="") as handle:
                writer = csv.DictWriter(handle, fieldnames=["cpu_id", "role", "ffmpeg_cpu_seconds"])
                writer.writeheader()
                writer.writerows([
                    {"cpu_id": "0", "role": "encoder", "ffmpeg_cpu_seconds": "10"},
                    {"cpu_id": "1", "role": "decoder", "ffmpeg_cpu_seconds": "2"},
                ])

            rows = scan(root, 59.5)
            self.assertEqual(rows[0]["noisy_neighbor"], "")
            self.assertEqual(rows[0]["noisy_neighbor_profile"], "")
            self.assertEqual(rows[0]["noisy_neighbor_image"], "")
            self.assertEqual(rows[0]["noisy_neighbor_args"], "")
            self.assertEqual(rows[0]["encoder_real_cpu_total_usage_avg_pct"], 10.0)
            self.assertEqual(rows[0]["decoder_real_cpu_total_usage_avg_pct"], 40.0)
            self.assertEqual(rows[0]["cross_numa_upi_bandwidth_gbps"], 2.5)
            self.assertEqual(rows[0]["cross_numa_upi_gbps_per_stream"], 2.5)

            html_path, xlsx_path = build_summary(root, 59.5)
            page = html_path.read_text()
            self.assertIn("Benchmark results — no noise", page)
            self.assertIn("Noisy-neighbor benchmark results", page)
            with (root / "summary.csv").open(newline="") as handle:
                header = next(csv.reader(handle))
            self.assertEqual(
                header[-3:],
                ["noisy_neighbor_image", "noisy_neighbor_args", "result_directory"],
            )
            self.assertIn("<dt>Encoder-CPU real usage</dt><dd>10.000%</dd>", page)
            self.assertIn("<dt>Decoder-CPU real usage</dt><dd>40.000%</dd>", page)
            self.assertIn("<dt>Cross-socket UPI incoming</dt><dd>2.5 GB/s</dd>", page)
            self.assertIn("<dt>UPI incoming per stream</dt><dd>2.5 GB/s</dd>", page)
            self.assertLess(
                page.index("Encoder-CPU real total usage<br>(avg %)"),
                page.index("Encoder avg utilization<br>per used CPU (%)"),
            )
            self.assertLess(
                page.index("Decoder-CPU real total usage<br>(avg %)"),
                page.index("Decoder avg utilization<br>per used CPU (%)"),
            )
            # ENC_CORES counts physical cores, so neither the column nor the legend
            # may call it a CPU count, and the report may not assume SMT is off.
            self.assertIn("Encoder cores<br>per session", page)
            self.assertNotIn("Encoder CPUs<br>per session", page)
            self.assertNotIn("no-SMT", page)
            self.assertTrue(xlsx_path.is_file())
            workbook = load_workbook(xlsx_path, read_only=True)
            self.assertIn("No noise", workbook.sheetnames)
            self.assertIn("Noisy neighbor", workbook.sheetnames)
            self.assertIn("Noise profile legend", workbook.sheetnames)
            self.assertIn("Noisy-neighbor scenario reference", page)
            self.assertIn("host-a is not a Pod", page)

    def test_result_directory_from_before_the_rename_is_still_read(self):
        # core-system-usage.csv used to call the column real_core_total_usage_pct.
        # Re-summarizing such a run must still find the metric, not silently drop it.
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            result = root / "pre-rename-run"
            result.mkdir()
            (result / "config.json").write_text(json.dumps({
                "SCENARIO": "pinned", "PLACEMENT": "exclusive", "STREAMS": "1",
            }))
            with (result / "metrics.csv").open("w", newline="") as handle:
                writer = csv.DictWriter(handle, fieldnames=["category", "metric", "value", "role"])
                writer.writeheader()
                writer.writerow({
                    "category": "FFmpeg progress", "metric": "fps",
                    "value": "avg=60;min=60;max=60", "role": "encoder",
                })
            with (result / "core-system-usage.csv").open("w", newline="") as handle:
                writer = csv.DictWriter(handle, fieldnames=["cpu_id", "real_core_total_usage_pct"])
                writer.writeheader()
                writer.writerow({"cpu_id": "0", "real_core_total_usage_pct": "55"})
            with (result / "cpu-usage.csv").open("w", newline="") as handle:
                writer = csv.DictWriter(handle, fieldnames=["cpu_id", "role", "ffmpeg_cpu_seconds"])
                writer.writeheader()
                writer.writerow({"cpu_id": "0", "role": "encoder", "ffmpeg_cpu_seconds": "10"})

            self.assertEqual(scan(root, 59.5)[0]["encoder_real_cpu_total_usage_avg_pct"], 55.0)

    def test_memory_metrics_are_summarized_per_run(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            result = root / "memory-run"
            result.mkdir()
            (result / "config.json").write_text(json.dumps({
                "SCENARIO": "pinned", "PLACEMENT": "exclusive", "STREAMS": "2",
            }))
            with (result / "metrics.csv").open("w", newline="") as handle:
                writer = csv.DictWriter(handle, fieldnames=["category", "metric", "value", "scope", "role"])
                writer.writeheader()
                writer.writerows([
                    {"category": "FFmpeg progress", "metric": "fps", "value": "avg=60;min=60;max=60", "scope": "", "role": "encoder"},
                    {"category": "Prometheus", "metric": "l3_cache_hit_ratio", "value": "avg=0.4;min=0.3;max=0.5", "scope": "{}", "role": ""},
                    {"category": "Prometheus", "metric": "l3_cache_misses_per_second", "value": "avg=250000000;min=1;max=2", "scope": "{}", "role": ""},
                    {"category": "Prometheus", "metric": "dram_read_bytes_per_second", "value": "avg=140000000000;min=1;max=2", "scope": "{}", "role": ""},
                    {"category": "Prometheus", "metric": "dram_write_bytes_per_second", "value": "avg=60000000000;min=1;max=2", "scope": "{}", "role": ""},
                ])

            row = scan(root, 59.5)[0]
            self.assertEqual(row["l3_cache_hit_ratio_pct"], 40.0)
            self.assertEqual(row["l3_cache_misses_per_second_millions"], 250.0)
            self.assertEqual(row["dram_read_gbps"], 140.0)
            self.assertEqual(row["dram_write_gbps"], 60.0)
            self.assertEqual(row["dram_total_gbps"], 200.0)
            self.assertEqual(row["dram_total_gbps_per_stream"], 100.0)

            page = build_summary(root, 59.5)[0].read_text()
            self.assertIn("Cache and DRAM traffic", page)
            self.assertIn("<td class='numeric'>40.0%</td>", page)
            self.assertIn("<td class='numeric'>200.0</td>", page)

    def test_memory_metrics_missing_are_unavailable(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            result = root / "no-pcm-run"
            result.mkdir()
            (result / "config.json").write_text(json.dumps({"SCENARIO": "pinned", "STREAMS": "1"}))
            with (result / "metrics.csv").open("w", newline="") as handle:
                writer = csv.DictWriter(handle, fieldnames=["category", "metric", "value", "role"])
                writer.writeheader()
                writer.writerow({
                    "category": "FFmpeg progress", "metric": "fps",
                    "value": "avg=60;min=60;max=60", "role": "encoder",
                })

            row = scan(root, 59.5)[0]
            for field in ("l3_cache_hit_ratio_pct", "dram_read_gbps", "dram_write_gbps", "dram_total_gbps"):
                self.assertEqual(row[field], "unavailable")

    def test_platform_spec_drives_peak_columns_and_panel(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            result = root / "platform-run"
            result.mkdir()
            (result / "config.json").write_text(json.dumps({
                "SCENARIO": "pinned", "STREAMS": "2", "NODE": "worker-1",
            }))
            (result / "host.json").write_text(json.dumps({"platform_spec": {
                "cpu_model": "Intel(R) Xeon(R) 6767P", "sockets": 2, "cores_per_socket": 64,
                "logical_cpus": 128, "populated_channels": 16, "memory_transfer_mt_s": 6400,
                "theoretical_dram_gbps": 819.2,
            }}))
            with (result / "metrics.csv").open("w", newline="") as handle:
                writer = csv.DictWriter(handle, fieldnames=["category", "metric", "value", "scope", "role"])
                writer.writeheader()
                writer.writerows([
                    {"category": "FFmpeg progress", "metric": "fps", "value": "avg=60;min=60;max=60", "scope": "", "role": "encoder"},
                    {"category": "Prometheus", "metric": "dram_read_bytes_per_second", "value": "avg=140000000000;min=1;max=2", "scope": "{}", "role": ""},
                    {"category": "Prometheus", "metric": "dram_write_bytes_per_second", "value": "avg=64960000000;min=1;max=2", "scope": "{}", "role": ""},
                ])

            row = scan(root, 59.5)[0]
            self.assertEqual(row["theoretical_dram_peak_gbps"], 819.2)
            self.assertEqual(row["dram_pct_of_theoretical_peak"], 25.02)

            page = build_summary(root, 59.5)[0].read_text()
            self.assertIn("Tested worker platform specification", page)
            self.assertIn("Intel(R) Xeon(R) 6767P", page)
            self.assertIn("<td class='numeric'>25.02%</td>", page)

    def test_platform_panel_reports_missing_spec(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            result = root / "old-run"
            result.mkdir()
            (result / "config.json").write_text(json.dumps({"SCENARIO": "pinned", "STREAMS": "1"}))
            with (result / "metrics.csv").open("w", newline="") as handle:
                writer = csv.DictWriter(handle, fieldnames=["category", "metric", "value", "role"])
                writer.writeheader()
                writer.writerow({
                    "category": "FFmpeg progress", "metric": "fps",
                    "value": "avg=60;min=60;max=60", "role": "encoder",
                })

            row = scan(root, 59.5)[0]
            self.assertEqual(row["theoretical_dram_peak_gbps"], "unavailable")
            self.assertEqual(row["dram_pct_of_theoretical_peak"], "unavailable")
            page = build_summary(root, 59.5)[0].read_text()
            self.assertIn("No platform specification captured", page)

    def test_rdt_monitoring_reports_both_groups(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            result = root / "rdt-monitor-run"
            result.mkdir()
            (result / "config.json").write_text(json.dumps({
                "SCENARIO": "pinned", "STREAMS": "1", "RDT_MONITOR": "1",
            }))
            (result / "rdt-summary.json").write_text(json.dumps([
                {"group": "encoder", "domain": "mon_L3_00",
                 "llc_occupancy_bytes_avg": 100 * 2**20, "mbm_total_bytes_per_second_avg": 4e9},
                {"group": "decoder", "domain": "mon_L3_00",
                 "llc_occupancy_bytes_avg": 50 * 2**20, "mbm_total_bytes_per_second_avg": 2e9},
                {"group": "noise", "domain": "mon_L3_00",
                 "llc_occupancy_bytes_avg": 200 * 2**20, "mbm_total_bytes_per_second_avg": 30e9},
            ]))
            with (result / "metrics.csv").open("w", newline="") as handle:
                writer = csv.DictWriter(handle, fieldnames=["category", "metric", "value", "role"])
                writer.writeheader()
                writer.writerow({
                    "category": "FFmpeg progress", "metric": "fps",
                    "value": "avg=60;min=60;max=60", "role": "encoder",
                })

            row = scan(root, 59.5)[0]
            self.assertEqual(row["rdt_encoder_llc_occupancy_mib"], 100.0)
            self.assertEqual(row["rdt_encoder_total_mbm_gbps"], 4.0)
            self.assertEqual(row["rdt_decoder_llc_occupancy_mib"], 50.0)
            self.assertEqual(row["rdt_decoder_total_mbm_gbps"], 2.0)
            self.assertEqual(row["rdt_workload_llc_occupancy_mib"], 150.0)
            self.assertEqual(row["rdt_workload_total_mbm_gbps"], 6.0)
            self.assertEqual(row["rdt_noise_llc_occupancy_mib"], 200.0)
            self.assertEqual(row["rdt_noise_total_mbm_gbps"], 30.0)

            page = build_summary(root, 59.5)[0].read_text()
            self.assertIn("RDT noise LLC occupancy (MiB)", page)
            self.assertIn("RDT encoder LLC occupancy (MiB)", page)
            self.assertIn("RDT decoder-group MBM total (GB/s)", page)

    def test_rdt_groups_unavailable_without_monitoring(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            result = root / "plain-run"
            result.mkdir()
            (result / "config.json").write_text(json.dumps({"SCENARIO": "pinned", "STREAMS": "1"}))
            with (result / "metrics.csv").open("w", newline="") as handle:
                writer = csv.DictWriter(handle, fieldnames=["category", "metric", "value", "role"])
                writer.writeheader()
                writer.writerow({
                    "category": "FFmpeg progress", "metric": "fps",
                    "value": "avg=60;min=60;max=60", "role": "encoder",
                })

            row = scan(root, 59.5)[0]
            for field in ("rdt_workload_llc_occupancy_mib", "rdt_noise_llc_occupancy_mib",
                          "rdt_workload_total_mbm_gbps", "rdt_noise_total_mbm_gbps"):
                self.assertEqual(row[field], "unavailable")

    def test_rdt_control_profiles_are_explained(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            result = root / "rdt-run"
            result.mkdir()
            (result / "config.json").write_text(json.dumps({
                "SCENARIO": "pinned", "STREAMS": "1", "RDT_CONTROL_PROFILE": "mba-10",
            }))
            with (result / "metrics.csv").open("w", newline="") as handle:
                writer = csv.DictWriter(handle, fieldnames=["category", "metric", "value", "role"])
                writer.writeheader()
                writer.writerow({
                    "category": "FFmpeg progress", "metric": "fps",
                    "value": "avg=60;min=60;max=60", "role": "encoder",
                })

            html_path, xlsx_path = build_summary(root, 59.5)
            page = html_path.read_text()
            self.assertIn("RDT control profile reference", page)
            for profile in ("none", "cat-guarded", "cat-strong", "cat-16-1",
                            "mba-80", "mba-60 / mba-40 / mba-10", "mba-20",
                            "&lt;cat profile&gt;+mba-&lt;level&gt;"):
                self.assertIn(f"<strong>{profile}</strong>", page)
            self.assertIn("Memory Bandwidth Allocation (MBA)", page)
            self.assertIn("L3 Cache Allocation Technology (CAT)", page)
            self.assertIn("How to verify it worked", page)
            self.assertIn("Applied schemata", page)
            self.assertIn("How a profile is applied", page)
            self.assertIn("L3 mask fff0", page)
            self.assertIn("MB 80 for the noise group", page)
            workbook = load_workbook(xlsx_path, read_only=True)
            self.assertIn("RDT profile legend", workbook.sheetnames)
            header = [cell.value for cell in next(workbook["RDT profile legend"].iter_rows(max_row=1))]
            self.assertEqual(header[2], "Applied schemata")
            self.assertEqual(header[5], "How to verify")

    def test_noisy_neighbor_requires_running_evidence(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            result = root / "noise-run"
            result.mkdir()
            config = {
                "SCENARIO": "baseline",
                "PLACEMENT": "free",
                "STREAMS": "1",
                "NOISY_NEIGHBOR_ENABLED": "1",
                "NOISY_NEIGHBOR_PROFILE": "pod-a",
                "NOISY_NEIGHBOR_IMAGE": "stress-ng:test",
                "NOISY_NEIGHBOR_ARGS": "--cpu 4",
            }
            (result / "config.json").write_text(json.dumps(config))
            with (result / "metrics.csv").open("w", newline="") as handle:
                writer = csv.DictWriter(handle, fieldnames=["category", "metric", "value", "role"])
                writer.writeheader()
                writer.writerow({
                    "category": "FFmpeg progress", "metric": "fps",
                    "value": "avg=60;min=60;max=60", "role": "encoder",
                })

            self.assertEqual(scan(root, 59.5)[0]["noisy_neighbor"], "")
            (result / "noisy-neighbor-after.json").write_text(json.dumps({"phase": "Running"}))
            row = scan(root, 59.5)[0]
            self.assertEqual(row["noisy_neighbor"], "enabled")
            self.assertEqual(row["noisy_neighbor_profile"], "pod-a")
            self.assertEqual(row["noisy_neighbor_image"], "stress-ng:test")
            self.assertEqual(row["noisy_neighbor_numa_status"], "n/a")
            self.assertEqual(row["noisy_neighbor_same_numa"], "n/a")
            self.assertEqual(row["noisy_neighbor_precheck_status"], "n/a")

            self.assertEqual(row["noisy_neighbor_args"], "--cpu 4")
            self.assertEqual(row["cross_numa_upi_bandwidth_gbps"], "unavailable")
            self.assertEqual(row["cross_numa_upi_gbps_per_stream"], "unavailable")

            (result / "noisy-neighbor-after.json").write_text(json.dumps({
                "phase": "Running",
                "numa": {"status": "same-socket", "same_socket": True, "precheck": {"status": "same-socket"}},
            }))
            row = scan(root, 59.5)[0]
            self.assertEqual(row["noisy_neighbor_numa_status"], "same-socket")
            self.assertEqual(row["noisy_neighbor_same_numa"], "yes")
            self.assertEqual(row["noisy_neighbor_precheck_status"], "same-socket")


            page = build_summary(root, 59.5)[0].read_text()
            self.assertIn("<p>No passing scenarios yet.</p>", page)
            self.assertIn("<td class='numeric'>unavailable</td>", page)
            self.assertNotIn("unavailable%", page)

    def test_best_scenario_link_uses_clean_run_not_noisy_neighbor(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)

            def make_result(name: str, noisy: bool, encoder_cpu: float) -> None:
                result = root / name
                result.mkdir()
                config = {
                    "SCENARIO": "pinned",
                    "PLACEMENT": "exclusive",
                    "STREAMS": "20",
                    "PRESET": "medium",
                    "ENC_THREADS": "15",
                    "SLICES": "2",
                    "NOISY_NEIGHBOR_ENABLED": "1" if noisy else "0",
                    "NOISY_NEIGHBOR_PROFILE": "pod-b" if noisy else "",
                }
                (result / "config.json").write_text(json.dumps(config))
                with (result / "metrics.csv").open("w", newline="") as handle:
                    writer = csv.DictWriter(handle, fieldnames=["category", "metric", "value", "role"])
                    writer.writeheader()
                    writer.writerows([
                        {"category": "FFmpeg progress", "metric": "fps", "value": "avg=60;min=60;max=60", "role": "encoder"},
                        {"category": "CPU", "metric": "FFmpeg CPU demand", "value": str(encoder_cpu), "role": "encoder"},
                        {"category": "CPU", "metric": "FFmpeg CPU demand", "value": "8", "role": "decoder"},
                    ])
                if noisy:
                    (result / "noisy-neighbor-after.json").write_text(json.dumps({"phase": "Running"}))

            make_result("clean-best", noisy=False, encoder_cpu=90)
            make_result("noisy-lower-cpu", noisy=True, encoder_cpu=80)

            html_path, xlsx_path = build_summary(root, 59.5)
            page = html_path.read_text()
            winner_section = page[page.index("Best result by scenario"):page.index("Benchmark results — no noise")]
            self.assertIn("clean-best/report.html", winner_section)
            self.assertNotIn("noisy-lower-cpu/report.html", winner_section)

            workbook = load_workbook(xlsx_path, read_only=True)
            winner_rows = list(workbook["Best by scenario"].iter_rows(values_only=True))
            self.assertTrue(str(winner_rows[1][-1]).endswith("clean-best"))


if __name__ == "__main__":
    unittest.main()
