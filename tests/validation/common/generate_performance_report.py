#!/usr/bin/env python3
"""
Generate XLSX performance report from pytest log files.

Parses pytest.log files produced by dual-host VF multisession FPS tests
and generates a formatted Excel workbook with:
  - Summary sheet: SW/HW configuration tables + pass/fail matrix
  - Data sheets (one per test type): TX Only, RX Only, Both RX TX,
    TX Redundant, RX Redundant, Both RX TX Redundant

Each data sheet contains per-session FPS, frames, and throughput data
for both DSA and non-DSA test variants.

Usage:
    python generate_performance_report.py <path> [-o output.xlsx] [-p platform_config.json]
    python generate_performance_report.py logs/performance/latest/pytest.log
    python generate_performance_report.py logs/performance/

The <path> can be:
  - A single pytest.log file
  - A directory (searched recursively for all pytest.log files)

When a directory is given, all pytest.log files found in subdirectories
are parsed and merged. If the same test case appears in multiple log files,
only the latest result is kept (based on the timestamp in the directory name).

Use -p/--platform-config to provide a JSON file with SW/HW platform details.
See platform_config.json for the expected format.

If -o is not given, the output file will be placed next to the input path
as performance_report_<timestamp>.xlsx.
"""

import argparse
import glob
import json
import os
import re
import sys
from collections import OrderedDict
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Alignment,
        Border,
        Font,
        PatternFill,
        Side,
    )
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl is required. Install it with: pip install openpyxl")
    sys.exit(1)


# ──────────────────────────────────────────────────────────────────────
# Data classes
# ──────────────────────────────────────────────────────────────────────

@dataclass
class SessionResult:
    """Per-session measurement data."""

    session_id: int = 0
    fps_requested: float = 0.0
    fps_avg: float = 0.0
    fps_min: float = 0.0
    fps_max: float = 0.0
    fps_pass: bool = False
    frames_tx: int = 0
    frames_rx: int = 0
    frame_success_pct: float = 0.0
    throughput_avg: float = 0.0
    throughput_min: float = 0.0
    throughput_max: float = 0.0
    throughput_direction: str = ""


@dataclass
class TestResult:
    """Complete result for one test parametrize variant."""

    test_name: str = ""
    test_func: str = ""
    direction: str = ""           # TX or RX
    test_label: str = ""          # Single-Core, Multi-Core, REDUNDANT ...
    num_sessions: int = 0
    target_fps: float = 0.0
    passed_sessions: int = 0
    total_sessions: int = 0
    use_dsa: bool = False
    dsa_device: str = ""
    overall_pass: bool = False
    # Config info
    media_file: str = ""
    media_format: str = ""
    resolution: str = ""
    measured_host: str = ""
    measured_role: str = ""
    measured_nic: str = ""
    companion_host: str = ""
    companion_role: str = ""
    companion_nic: str = ""
    nic_numa: str = ""
    nic_type: str = ""            # PF or VF
    dsa_numa: str = ""
    numa_match: str = ""
    redundant: bool = False
    multi_core: bool = False
    # Per-session data
    sessions: List[SessionResult] = field(default_factory=list)
    # DEV rates
    dev_rate_tx: float = 0.0
    dev_rate_rx: float = 0.0
    companion_dev_rate_tx: float = 0.0
    companion_dev_rate_rx: float = 0.0
    # Throughput per session (both measured and companion)
    measured_throughput: List[SessionResult] = field(default_factory=list)
    companion_throughput: List[SessionResult] = field(default_factory=list)


# ──────────────────────────────────────────────────────────────────────
# Regex patterns for parsing pytest.log
# ──────────────────────────────────────────────────────────────────────

# Test identification
RE_TEST_START = re.compile(
    r"Test (tests/.*?test_vf_multisession_fps_variants_dualhost\.py::"
    r"(test_dualhost_vf_\w+)\[(\d+)sess-(\d+)fps-(dsa|no-dsa)\]) was implemented"
)

# Configuration summary box fields
RE_CFG_TEST = re.compile(r"║\s+Test:\s+(.+?)\s*║")
RE_CFG_SESSIONS = re.compile(r"║\s+Sessions:\s+(\d+)\s+║")
RE_CFG_FPS = re.compile(r"║\s+Target FPS:\s+([\d.]+)\s+║")
RE_CFG_MEDIA = re.compile(r"║\s+Media:\s+(\S+)\s+║")
RE_CFG_FORMAT = re.compile(r"║\s+Format:\s+(\S+)\s+║")
RE_CFG_RESOLUTION = re.compile(r"║\s+Resolution:\s+(\S+)\s+║")
RE_CFG_HOST = re.compile(r"║\s+Host:\s+(\S+)\s+║")
RE_CFG_ROLE = re.compile(r"║\s+Role:\s+(\S+)\s+║")
RE_CFG_NIC = re.compile(r"║\s+NIC \(primary\):\s+(\S+)\s+║")
RE_CFG_NIC_NUMA = re.compile(r"║\s+NIC NUMA:\s+(\S+)\s+║")
RE_CFG_DSA = re.compile(r"║\s+DSA:\s+(.+?)\s*║")
RE_CFG_DSA_NUMA = re.compile(r"║\s+DSA NUMA:\s+(\S+)\s+║")
RE_CFG_NUMA_MATCH = re.compile(r"║\s+NUMA Match:\s+(.+?)\s*║")

# Results header
RE_RESULTS_HEADER = re.compile(
    r"(TX|RX) Results\s*(.*?):\s*(\d+)/(\d+) sessions at target (\d+) fps"
)

# Per-session FPS line
RE_SESSION_FPS = re.compile(
    r"Session (\d+): FPS: requested=(\d+), avg=([\d.]+), min=([\d.]+), max=([\d.]+) ([✓✗])"
    r" \| Frames: TX=(\d+), RX=(\d+), Success=([\d.]+)%"
)

# Throughput section headers
RE_THROUGHPUT_MEASURED = re.compile(
    r"THROUGHPUT - Measured App \((TX|RX)\)"
)
RE_THROUGHPUT_COMPANION = re.compile(
    r"THROUGHPUT - Companion App \((TX|RX)\)"
)

# DEV rate lines
RE_DEV_RATE = re.compile(
    r"DEV Avr rate (TX|RX): ([\d.]+) Mb/s"
)

# Per-session throughput
RE_SESSION_THROUGHPUT = re.compile(
    r"Session (\d+) \((TX|RX)\): throughput avg=([\d.]+), min=([\d.]+), max=([\d.]+) Mb/s"
)

# Pass/fail outcome
RE_TEST_PASS = re.compile(
    r"TEST_PASS\s+.*?passed: (\d+) sessions @ (\d+) fps"
)
RE_TEST_FAIL_ONLY = re.compile(
    r"TEST_FAIL\s+Only (\d+)/(\d+) sessions reached target (\d+) fps"
)
RE_TEST_PASS_LOG = re.compile(
    r"TEST_PASS\s+Test passed for .*?::(test_dualhost_vf_\w+)\[(\d+)sess-(\d+)fps-(dsa|no-dsa)\]"
)
RE_TEST_FAIL_LOG = re.compile(
    r"TEST_FAIL\s+Test failed for .*?::(test_dualhost_vf_\w+)\[(\d+)sess-(\d+)fps-(dsa|no-dsa)\]"
)


# ──────────────────────────────────────────────────────────────────────
# Sheet category mapping
# ──────────────────────────────────────────────────────────────────────

# Map test function names to sheet categories
FUNC_TO_CATEGORY = {
    "test_dualhost_vf_tx_fps_variants_single_core": "TX Only",
    "test_dualhost_vf_tx_fps_variants_multi_core": "TX Only",
    "test_dualhost_vf_rx_fps_variants_single_core": "RX Only",
    "test_dualhost_vf_rx_fps_variants_multi_core": "RX Only",
    "test_dualhost_vf_tx_fps_variants_single_core_redundant": "TX Redundant",
    "test_dualhost_vf_tx_fps_variants_multi_core_redundant": "TX Redundant",
    "test_dualhost_vf_rx_fps_variants_single_core_redundant": "RX Redundant",
    "test_dualhost_vf_rx_fps_variants_multi_core_redundant": "RX Redundant",
}

# The 6 data sheet names in order
DATA_SHEET_NAMES = [
    "TX Only",
    "RX Only",
    "Both RX TX",
    "TX Redundant",
    "RX Redundant",
    "Both RX TX Redundant",
]


# ──────────────────────────────────────────────────────────────────────
# Parser
# ──────────────────────────────────────────────────────────────────────

def parse_pytest_log(log_path: str) -> List[TestResult]:
    """Parse a pytest.log file and extract all test results."""
    results: List[TestResult] = []

    with open(log_path, "r", encoding="utf-8", errors="replace") as f:
        lines = f.readlines()

    current: Optional[TestResult] = None
    in_config_box = False
    config_section = ""  # "measured" or "companion"
    in_throughput_measured = False
    in_throughput_companion = False

    for i, line in enumerate(lines):
        # ── Test start ──
        m = RE_TEST_START.search(line)
        if m:
            # Save previous result if any
            if current is not None:
                results.append(current)

            current = TestResult()
            current.test_name = m.group(1)
            current.test_func = m.group(2)
            current.num_sessions = int(m.group(3))
            current.target_fps = float(m.group(4))
            current.use_dsa = m.group(5) == "dsa"

            # Derive properties from function name
            if "redundant" in current.test_func:
                current.redundant = True
            if "multi_core" in current.test_func:
                current.multi_core = True
            if "_tx_" in current.test_func:
                current.direction = "TX"
            elif "_rx_" in current.test_func:
                current.direction = "RX"

            # Detect NIC interface type from function name
            if "_vf_" in current.test_func:
                current.nic_type = "VF"
            else:
                current.nic_type = "PF"

            core_mode = "Multi-Core" if current.multi_core else "Single-Core"
            mode = "REDUNDANT " if current.redundant else ""
            current.test_label = f"{mode}{current.direction} {core_mode}"

            in_config_box = False
            config_section = ""
            in_throughput_measured = False
            in_throughput_companion = False
            continue

        if current is None:
            continue

        # ── Configuration summary box ──
        if "TEST CONFIGURATION SUMMARY" in line:
            in_config_box = True
            config_section = ""
            continue

        if in_config_box:
            if "╚" in line:
                in_config_box = False
                continue

            if "MEASURED HOST" in line:
                config_section = "measured"
                continue
            if "COMPANION HOST" in line:
                config_section = "companion"
                continue

            m = RE_CFG_MEDIA.search(line)
            if m:
                current.media_file = m.group(1).strip()
            m = RE_CFG_FORMAT.search(line)
            if m:
                current.media_format = m.group(1).strip()
            m = RE_CFG_RESOLUTION.search(line)
            if m:
                current.resolution = m.group(1).strip()

            m = RE_CFG_HOST.search(line)
            if m:
                if config_section == "measured":
                    current.measured_host = m.group(1).strip()
                elif config_section == "companion":
                    current.companion_host = m.group(1).strip()

            m = RE_CFG_ROLE.search(line)
            if m:
                if config_section == "measured":
                    current.measured_role = m.group(1).strip()
                elif config_section == "companion":
                    current.companion_role = m.group(1).strip()

            m = RE_CFG_NIC.search(line)
            if m:
                if config_section == "measured":
                    current.measured_nic = m.group(1).strip()
                elif config_section == "companion":
                    current.companion_nic = m.group(1).strip()

            m = RE_CFG_NIC_NUMA.search(line)
            if m and config_section == "measured":
                current.nic_numa = m.group(1).strip()

            m = RE_CFG_DSA.search(line)
            if m:
                dsa_text = m.group(1).strip()
                if "ENABLED" in dsa_text:
                    dm = re.search(r"\(([^)]+)\)", dsa_text)
                    if dm:
                        current.dsa_device = dm.group(1)

            m = RE_CFG_DSA_NUMA.search(line)
            if m:
                current.dsa_numa = m.group(1).strip()
            m = RE_CFG_NUMA_MATCH.search(line)
            if m:
                current.numa_match = m.group(1).strip()
            continue

        # ── Results header ──
        m = RE_RESULTS_HEADER.search(line)
        if m:
            current.direction = m.group(1)
            dsa_text = m.group(2).strip()
            current.passed_sessions = int(m.group(3))
            current.total_sessions = int(m.group(4))
            current.target_fps = float(m.group(5))
            if "with DSA" in dsa_text:
                current.use_dsa = True
            in_throughput_measured = False
            in_throughput_companion = False
            continue

        # ── Per-session FPS ──
        m = RE_SESSION_FPS.search(line)
        if m and not in_throughput_measured and not in_throughput_companion:
            sr = SessionResult(
                session_id=int(m.group(1)),
                fps_requested=float(m.group(2)),
                fps_avg=float(m.group(3)),
                fps_min=float(m.group(4)),
                fps_max=float(m.group(5)),
                fps_pass=(m.group(6) == "✓"),
                frames_tx=int(m.group(7)),
                frames_rx=int(m.group(8)),
                frame_success_pct=float(m.group(9)),
            )
            current.sessions.append(sr)
            continue

        # ── Throughput sections ──
        m = RE_THROUGHPUT_MEASURED.search(line)
        if m:
            in_throughput_measured = True
            in_throughput_companion = False
            continue

        m = RE_THROUGHPUT_COMPANION.search(line)
        if m:
            in_throughput_companion = True
            in_throughput_measured = False
            continue

        # ── DEV rate ──
        m = RE_DEV_RATE.search(line)
        if m:
            direction = m.group(1)
            rate = float(m.group(2))
            if in_throughput_measured:
                if direction == "TX":
                    current.dev_rate_tx = rate
                else:
                    current.dev_rate_rx = rate
            elif in_throughput_companion:
                if direction == "TX":
                    current.companion_dev_rate_tx = rate
                else:
                    current.companion_dev_rate_rx = rate
            continue

        # ── Per-session throughput ──
        m = RE_SESSION_THROUGHPUT.search(line)
        if m:
            tp = SessionResult(
                session_id=int(m.group(1)),
                throughput_direction=m.group(2),
                throughput_avg=float(m.group(3)),
                throughput_min=float(m.group(4)),
                throughput_max=float(m.group(5)),
            )
            if in_throughput_measured:
                current.measured_throughput.append(tp)
            elif in_throughput_companion:
                current.companion_throughput.append(tp)
            continue

        # ── Pass outcome (in-test log) ──
        m = RE_TEST_PASS.search(line)
        if m:
            current.overall_pass = True
            continue

        # ── Fail outcome (in-test log) ──
        m = RE_TEST_FAIL_ONLY.search(line)
        if m:
            current.overall_pass = False
            continue

        # ── Final pass/fail from pytest framework ──
        m = RE_TEST_PASS_LOG.search(line)
        if m and current and m.group(1) == current.test_func:
            current.overall_pass = True
            results.append(current)
            current = None
            continue

        m = RE_TEST_FAIL_LOG.search(line)
        if m and current and m.group(1) == current.test_func:
            current.overall_pass = False
            results.append(current)
            current = None
            continue

    # Don't forget the last one
    if current is not None:
        results.append(current)

    return results


# ──────────────────────────────────────────────────────────────────────
# Categorize results into sheets
# ──────────────────────────────────────────────────────────────────────

def categorize_results(
    results: List[TestResult],
) -> Dict[str, List[TestResult]]:
    """Sort test results into the 6 data sheet categories."""
    categories: Dict[str, List[TestResult]] = {name: [] for name in DATA_SHEET_NAMES}

    for r in results:
        cat = FUNC_TO_CATEGORY.get(r.test_func, "")
        if cat:
            categories[cat].append(r)

    # Build "Both RX TX" sheets from combined TX+RX results
    # (if we have both TX and RX results, we combine them)
    # The "Both" sheets are populated only if both TX and RX tests ran
    # with the same session count / fps / dsa combination
    _build_both_sheet(categories, "TX Only", "RX Only", "Both RX TX")
    _build_both_sheet(categories, "TX Redundant", "RX Redundant", "Both RX TX Redundant")

    return categories


def _build_both_sheet(
    categories: Dict[str, List[TestResult]],
    tx_key: str,
    rx_key: str,
    both_key: str,
) -> None:
    """Populate 'Both' sheet from matching TX and RX results."""
    tx_results = categories.get(tx_key, [])
    rx_results = categories.get(rx_key, [])

    # Group by (sessions, fps, dsa, multi_core)
    tx_map = {}
    for r in tx_results:
        key = (r.num_sessions, r.target_fps, r.use_dsa, r.multi_core)
        tx_map[key] = r

    for rx in rx_results:
        key = (rx.num_sessions, rx.target_fps, rx.use_dsa, rx.multi_core)
        if key in tx_map:
            # Create a combined result referencing both
            combined = TestResult(
                test_name=f"TX+RX: {tx_map[key].test_name} & {rx.test_name}",
                test_func=f"combined_{tx_map[key].test_func}_{rx.test_func}",
                direction="TX+RX",
                test_label=f"TX+RX {'Multi-Core' if rx.multi_core else 'Single-Core'}",
                num_sessions=rx.num_sessions,
                target_fps=rx.target_fps,
                use_dsa=rx.use_dsa,
                dsa_device=rx.dsa_device or tx_map[key].dsa_device,
                overall_pass=tx_map[key].overall_pass and rx.overall_pass,
                media_file=rx.media_file,
                media_format=rx.media_format,
                resolution=rx.resolution,
                measured_host=rx.measured_host,
                measured_nic=rx.measured_nic,
                companion_host=rx.companion_host,
                companion_nic=rx.companion_nic,
                nic_type=rx.nic_type or tx_map[key].nic_type,
                redundant=rx.redundant,
                multi_core=rx.multi_core,
            )
            # Combine sessions: TX from tx_map, RX from rx
            combined.sessions = tx_map[key].sessions + rx.sessions
            categories[both_key].append(combined)


# ──────────────────────────────────────────────────────────────────────
# Styles
# ──────────────────────────────────────────────────────────────────────

# Colors
PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
SUBHEADER_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
SUBHEADER_FONT = Font(bold=True, size=10)
TITLE_FONT = Font(bold=True, size=14)
SECTION_FONT = Font(bold=True, size=12)
LABEL_FONT = Font(bold=True, size=10)
VALUE_FONT = Font(size=10)
PASS_FONT = Font(color="006100", bold=True)
FAIL_FONT = Font(color="9C0006", bold=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _apply_header_style(ws, row, col_start, col_end):
    """Apply header styling to a range of cells."""
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def _apply_subheader_style(ws, row, col_start, col_end):
    """Apply sub-header styling to a range of cells."""
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = SUBHEADER_FILL
        cell.font = SUBHEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def _write_cell(ws, row, col, value, font=None, fill=None, alignment=None, border=None):
    """Write a cell with optional styling."""
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border
    return cell


# ──────────────────────────────────────────────────────────────────────
# Summary sheet
# ──────────────────────────────────────────────────────────────────────

def _collect_system_info(results: List[TestResult]) -> Dict[str, str]:
    """Collect system info from test results to build the config tables."""
    info = {
        "Measured Host": "",
        "Measured NIC": "",
        "NIC NUMA": "",
        "NIC Interface Type": "",
        "Companion Host": "",
        "Companion NIC": "",
        "Media File": "",
        "Media Format": "",
        "Resolution": "",
        "DSA Device": "",
        "DSA NUMA": "",
        "NUMA Match": "",
    }
    for r in results:
        if r.measured_host and not info["Measured Host"]:
            info["Measured Host"] = r.measured_host
        if r.measured_nic and not info["Measured NIC"]:
            info["Measured NIC"] = r.measured_nic
        if r.nic_numa and r.nic_numa != "?" and not info["NIC NUMA"]:
            info["NIC NUMA"] = r.nic_numa
        if r.nic_type and not info["NIC Interface Type"]:
            info["NIC Interface Type"] = r.nic_type
        if r.companion_host and not info["Companion Host"]:
            info["Companion Host"] = r.companion_host
        if r.companion_nic and not info["Companion NIC"]:
            info["Companion NIC"] = r.companion_nic
        if r.media_file and not info["Media File"]:
            info["Media File"] = r.media_file
        if r.media_format and not info["Media Format"]:
            info["Media Format"] = r.media_format
        if r.resolution and not info["Resolution"]:
            info["Resolution"] = r.resolution
        if r.use_dsa and r.dsa_device and not info["DSA Device"]:
            info["DSA Device"] = r.dsa_device
        if r.dsa_numa and r.dsa_numa != "?" and not info["DSA NUMA"]:
            info["DSA NUMA"] = r.dsa_numa
        if r.numa_match and not info["NUMA Match"]:
            info["NUMA Match"] = r.numa_match
    return info


def _load_platform_config(config_path: Optional[str],
                         log_dir: Optional[str] = None) -> Dict[str, Any]:
    """Load platform configuration from a JSON file.

    Search order when config_path is None:
      1. platform_config.json inside the log directory (auto-collected during tests)
      2. platform_config.json in the same directory as this script (static fallback)
    """
    if config_path is None:
        # First check inside the log directory (auto-collected by tests)
        if log_dir:
            log_cfg = os.path.join(log_dir, "platform_config.json")
            if os.path.isfile(log_cfg):
                config_path = log_cfg
        # Fallback to script directory
        if config_path is None:
            script_dir = os.path.dirname(os.path.abspath(__file__))
            default_path = os.path.join(script_dir, "platform_config.json")
            if os.path.isfile(default_path):
                config_path = default_path
            else:
                return {}

    if not os.path.isfile(config_path):
        print(f"WARNING: Platform config not found: {config_path}")
        return {}

    try:
        with open(config_path, "r") as f:
            config = json.load(f)
        print(f"  Loaded platform config from: {config_path}")
        return config
    except (json.JSONDecodeError, IOError) as e:
        print(f"WARNING: Failed to parse platform config: {e}")
        return {}


def _write_config_table(ws, row, title, items, col_start=1):
    """Write a configuration table with Category | Ingredient | Version columns.

    Args:
        ws: Worksheet
        row: Starting row
        title: Table title
        items: List of (category, ingredient, version) tuples.
               category="" means it's a sub-item of the previous category.
        col_start: Starting column

    Returns:
        Next available row after the table.
    """
    # Title
    _write_cell(ws, row, col_start, title, font=SECTION_FONT)
    row += 1

    # Header
    headers = ["", "Ingredient", "Version"]
    for ci, h in enumerate(headers):
        _write_cell(ws, row, col_start + ci, h,
                    font=HEADER_FONT, fill=HEADER_FILL,
                    border=THIN_BORDER, alignment=CENTER)
    row += 1

    # Data rows
    for category, ingredient, version in items:
        _write_cell(ws, row, col_start, category, font=LABEL_FONT, border=THIN_BORDER, alignment=LEFT)
        _write_cell(ws, row, col_start + 1, ingredient, font=VALUE_FONT, border=THIN_BORDER, alignment=LEFT)
        _write_cell(ws, row, col_start + 2, version, font=VALUE_FONT, border=THIN_BORDER, alignment=WRAP)
        row += 1

    return row


def _write_simple_table(ws, row, title, items, col_start=1):
    """Write a simple 2-column table (Parameter | Value).

    Args:
        ws: Worksheet
        row: Starting row
        title: Table title
        items: List of (parameter, value) tuples
        col_start: Starting column

    Returns:
        Next available row after the table.
    """
    _write_cell(ws, row, col_start, title, font=SECTION_FONT)
    row += 1

    _write_cell(ws, row, col_start, "Parameter", font=HEADER_FONT, fill=HEADER_FILL, border=THIN_BORDER, alignment=CENTER)
    ws.merge_cells(start_row=row, start_column=col_start + 1,
                  end_row=row, end_column=col_start + 3)
    _write_cell(ws, row, col_start + 1, "Value", font=HEADER_FONT, fill=HEADER_FILL, border=THIN_BORDER, alignment=CENTER)
    row += 1

    for param, value in items:
        _write_cell(ws, row, col_start, param, font=LABEL_FONT, border=THIN_BORDER, alignment=LEFT)
        ws.merge_cells(start_row=row, start_column=col_start + 1,
                      end_row=row, end_column=col_start + 3)
        _write_cell(ws, row, col_start + 1, value, font=VALUE_FONT, border=THIN_BORDER, alignment=WRAP)
        row += 1

    return row


def _write_pass_fail_matrix(ws, row, title, results, all_fps):
    """Write a pass/fail matrix table for a set of results.

    Rows = session counts, Columns = FPS values (DSA / no-DSA).

    Args:
        ws: Worksheet
        row: Starting row
        title: Table title
        results: List of TestResult to include
        all_fps: Sorted list of all FPS values across all results

    Returns:
        Next available row after the table.
    """
    if not results:
        return row

    _write_cell(ws, row, 1, title, font=SECTION_FONT)
    row += 1

    cat_sessions = sorted(set(r.num_sessions for r in results))

    # Build lookup: (sessions, fps, use_dsa) -> overall_pass
    lookup = {}
    for r in results:
        key = (r.num_sessions, r.target_fps, r.use_dsa)
        lookup[key] = r.overall_pass

    # Header row
    col = 1
    _write_cell(ws, row, col, "Sessions", font=HEADER_FONT, fill=HEADER_FILL, border=THIN_BORDER, alignment=CENTER)
    col += 1
    for fps in all_fps:
        _write_cell(ws, row, col, f"{int(fps)} fps\nDSA", font=HEADER_FONT, fill=HEADER_FILL, border=THIN_BORDER,
                   alignment=Alignment(horizontal="center", vertical="center", wrap_text=True))
        col += 1
        _write_cell(ws, row, col, f"{int(fps)} fps\nno-DSA", font=HEADER_FONT, fill=HEADER_FILL, border=THIN_BORDER,
                   alignment=Alignment(horizontal="center", vertical="center", wrap_text=True))
        col += 1
    row += 1

    # Data rows
    for sess in cat_sessions:
        col = 1
        _write_cell(ws, row, col, sess, font=VALUE_FONT, border=THIN_BORDER, alignment=CENTER)
        col += 1
        for fps in all_fps:
            for dsa in [True, False]:
                key = (sess, fps, dsa)
                if key in lookup:
                    status = "PASS" if lookup[key] else "FAIL"
                    fill = PASS_FILL if lookup[key] else FAIL_FILL
                    font = PASS_FONT if lookup[key] else FAIL_FONT
                    _write_cell(ws, row, col, status, font=font, fill=fill, border=THIN_BORDER, alignment=CENTER)
                else:
                    _write_cell(ws, row, col, "-", font=VALUE_FONT, border=THIN_BORDER, alignment=CENTER)
                col += 1
        row += 1

    return row + 1


def _auto_adjust_columns(ws, min_width: int = 8, max_width: int = 60, padding: int = 3) -> None:
    """Auto-adjust column widths to fit cell content.

    Args:
        ws: Worksheet to adjust.
        min_width: Minimum column width.
        max_width: Maximum column width.
        padding: Extra characters of padding to add.
    """
    for col_cells in ws.columns:
        max_len = min_width
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is None:
                continue
            text = str(cell.value)
            # For multi-line cells, take the longest line
            lines = text.split("\n")
            cell_len = max(len(line) for line in lines) + padding
            # Bold text needs a little extra space
            if cell.font and cell.font.bold:
                cell_len += 2
            if cell_len > max_len:
                max_len = cell_len
        ws.column_dimensions[col_letter].width = min(max_len, max_width)


def write_summary_sheet(
    wb: Workbook,
    results: List[TestResult],
    categories: Dict[str, List[TestResult]],
    log_path: str,
    platform_config: Optional[Dict[str, Any]] = None,
) -> None:
    """Write the Summary sheet with SW/HW config tables and pass/fail matrices."""
    ws = wb.active
    ws.title = "Summary"

    if platform_config is None:
        platform_config = {}

    sw_cfg = platform_config.get("sw_configuration", {})
    hw_cfg = platform_config.get("hw_configuration", {})

    row = 1

    # ── Title ──
    _write_cell(ws, row, 1, "MTL Performance Test Report", font=TITLE_FONT)
    row += 1
    _write_cell(ws, row, 1, f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", font=VALUE_FONT)
    row += 1
    _write_cell(ws, row, 1, f"Source: {log_path}", font=VALUE_FONT)
    row += 2

    # ── SW Configuration Table ──
    sw_items = [
        ("Operating system", "OS", sw_cfg.get("os", "")),
        ("", "kernel", sw_cfg.get("kernel", "")),
        ("Package components", "MTL version", sw_cfg.get("mtl_version", "")),
        ("Dependency", "DPDK Driver", sw_cfg.get("dpdk_driver", "")),
        ("", "ICE Version", sw_cfg.get("ice_version", "")),
        ("", "DDP Version", sw_cfg.get("ddp_version", "")),
        ("", "BIOS VT-X & VT-D", sw_cfg.get("bios_vtx_vtd", "")),
        ("Others setting enforcing\nthe highest performance", "NIC ports settings in DPDK",
         sw_cfg.get("nic_ports_settings", "")),
        ("", "NIC ports order in tests configuration",
         sw_cfg.get("nic_ports_order", "")),
        ("", "Video files", sw_cfg.get("video_files", "")),
        ("", "Hugepages", sw_cfg.get("hugepages", "")),
        ("", "CPU cores", sw_cfg.get("cpu_cores", "")),
    ]
    row = _write_config_table(ws, row, "SW Configuration", sw_items)
    row += 1

    # ── HW Configuration Table ──
    hw_items = [
        ("Server", hw_cfg.get("server", "")),
        ("CPU", hw_cfg.get("cpu", "")),
        ("Memory", hw_cfg.get("memory", "")),
        ("NIC", hw_cfg.get("nic", "")),
    ]
    row = _write_simple_table(ws, row, "HW Configuration", hw_items)
    row += 1

    # ── NIC Configuration Table (from test results) ──
    sys_info = _collect_system_info(results)
    nic_items = [
        ("NIC Model", hw_cfg.get("nic", "")),
        ("NIC Interface Type", sys_info["NIC Interface Type"]),
        ("Measured NIC", sys_info["Measured NIC"]),
        ("NIC NUMA", sys_info["NIC NUMA"]),
        ("Measured Host", sys_info["Measured Host"]),
        ("Companion Host", sys_info["Companion Host"]),
        ("Companion NIC", sys_info["Companion NIC"]),
    ]
    row = _write_simple_table(ws, row, "NIC Configuration", nic_items)
    row += 1

    # ── DSA Configuration Table ──
    dsa_items = [
        ("DSA Device", sys_info["DSA Device"]),
        ("DSA NUMA", sys_info["DSA NUMA"]),
        ("NUMA Match", sys_info["NUMA Match"]),
    ]
    row = _write_simple_table(ws, row, "DSA Configuration", dsa_items)
    row += 1

    # ── Media Configuration ──
    media_items = [
        ("Media File", sys_info["Media File"]),
        ("Media Format", sys_info["Media Format"]),
        ("Resolution", sys_info["Resolution"]),
    ]
    row = _write_simple_table(ws, row, "Media Configuration", media_items)
    row += 1

    # ── Test Execution Summary ──
    total = len(results)
    passed = sum(1 for r in results if r.overall_pass)
    failed = total - passed
    exec_items = [
        ("Total Tests", str(total)),
        ("Passed", str(passed)),
        ("Failed", str(failed)),
        ("Pass Rate", f"{100 * passed / total:.1f}%" if total else "N/A"),
    ]
    _write_cell(ws, row, 1, "Test Execution Summary", font=SECTION_FONT)
    row += 1
    for k, v in exec_items:
        _write_cell(ws, row, 1, k, font=LABEL_FONT, border=THIN_BORDER, alignment=LEFT)
        cell = _write_cell(ws, row, 2, v, font=VALUE_FONT, border=THIN_BORDER, alignment=LEFT)
        if k == "Passed":
            cell.fill = PASS_FILL
        elif k == "Failed" and v != "0":
            cell.fill = FAIL_FILL
        row += 1
    row += 2

    # ── Pass/Fail Result Matrices — one table per category ──
    all_fps = sorted(set(r.target_fps for r in results))

    for category in DATA_SHEET_NAMES:
        cat_results = categories.get(category, [])
        if not cat_results:
            continue
        row = _write_pass_fail_matrix(ws, row, category, cat_results, all_fps)

    # ── Auto-adjust column widths ──
    _auto_adjust_columns(ws)


# ──────────────────────────────────────────────────────────────────────
# Data sheets
# ──────────────────────────────────────────────────────────────────────

def write_data_sheet(
    wb: Workbook,
    sheet_name: str,
    results: List[TestResult],
) -> None:
    """Write a data sheet with detailed per-session results."""
    if not results:
        return

    ws = wb.create_sheet(title=sheet_name)

    # Sort results by: use_dsa (DSA first), num_sessions, target_fps
    results_sorted = sorted(
        results,
        key=lambda r: (not r.use_dsa, r.num_sessions, r.target_fps),
    )

    row = 1

    # ── Sheet title ──
    _write_cell(ws, row, 1, f"{sheet_name} - Detailed Results", font=TITLE_FONT)
    row += 2

    for test_r in results_sorted:
        # ── Test header ──
        dsa_label = f" with DSA ({test_r.dsa_device})" if test_r.use_dsa else " without DSA"
        title = f"{test_r.test_label}{dsa_label} — {test_r.num_sessions} sessions @ {int(test_r.target_fps)} fps"
        _write_cell(ws, row, 1, title, font=SECTION_FONT)
        row += 1

        # Overall pass/fail
        status = "PASS" if test_r.overall_pass else "FAIL"
        status_fill = PASS_FILL if test_r.overall_pass else FAIL_FILL
        status_font = PASS_FONT if test_r.overall_pass else FAIL_FONT
        _write_cell(ws, row, 1, "Result:", font=LABEL_FONT)
        _write_cell(ws, row, 2, status, font=status_font, fill=status_fill, border=THIN_BORDER, alignment=CENTER)
        _write_cell(ws, row, 3, f"{test_r.passed_sessions}/{test_r.total_sessions} sessions passed",
                    font=VALUE_FONT)
        row += 2

        # ── Per-session FPS & Frames table ──
        if test_r.sessions:
            # Build throughput lookup: session_id -> avg throughput (Mb/s)
            tp_lookup = {}
            for tp in test_r.measured_throughput:
                tp_lookup[tp.session_id] = tp.throughput_avg

            headers = [
                "Session", "FPS Requested", "FPS Avg", "FPS Min", "FPS Max",
                "Pass", "Frames TX", "Frames RX", "Success %", "Avg Bitrate (Mb/s)",
            ]
            for ci, h in enumerate(headers, start=1):
                _write_cell(ws, row, ci, h, font=HEADER_FONT, fill=HEADER_FILL, border=THIN_BORDER, alignment=CENTER)
            row += 1

            for s in test_r.sessions:
                col = 1
                _write_cell(ws, row, col, s.session_id, font=VALUE_FONT, border=THIN_BORDER, alignment=CENTER)
                col += 1
                _write_cell(ws, row, col, int(s.fps_requested), font=VALUE_FONT, border=THIN_BORDER, alignment=CENTER)
                col += 1
                _write_cell(ws, row, col, round(s.fps_avg, 2), font=VALUE_FONT, border=THIN_BORDER, alignment=CENTER)
                col += 1
                _write_cell(ws, row, col, round(s.fps_min, 2), font=VALUE_FONT, border=THIN_BORDER, alignment=CENTER)
                col += 1
                _write_cell(ws, row, col, round(s.fps_max, 2), font=VALUE_FONT, border=THIN_BORDER, alignment=CENTER)
                col += 1
                pass_cell = _write_cell(
                    ws, row, col,
                    "✓" if s.fps_pass else "✗",
                    font=PASS_FONT if s.fps_pass else FAIL_FONT,
                    fill=PASS_FILL if s.fps_pass else FAIL_FILL,
                    border=THIN_BORDER,
                    alignment=CENTER,
                )
                col += 1
                _write_cell(ws, row, col, s.frames_tx, font=VALUE_FONT, border=THIN_BORDER, alignment=CENTER)
                col += 1
                _write_cell(ws, row, col, s.frames_rx, font=VALUE_FONT, border=THIN_BORDER, alignment=CENTER)
                col += 1
                _write_cell(ws, row, col, f"{s.frame_success_pct:.1f}%", font=VALUE_FONT, border=THIN_BORDER, alignment=CENTER)
                col += 1
                bitrate = tp_lookup.get(s.session_id)
                _write_cell(ws, row, col,
                            round(bitrate, 2) if bitrate is not None else "-",
                            font=VALUE_FONT, border=THIN_BORDER, alignment=CENTER)
                row += 1
            row += 1

        # ── Throughput summary ──
        has_measured_tp = test_r.measured_throughput or test_r.dev_rate_tx > 0 or test_r.dev_rate_rx > 0
        has_companion_tp = test_r.companion_throughput or test_r.companion_dev_rate_tx > 0 or test_r.companion_dev_rate_rx > 0

        if has_measured_tp:
            _write_cell(ws, row, 1, f"Measured App Throughput ({test_r.direction})", font=SUBHEADER_FONT, fill=SUBHEADER_FILL)
            row += 1
            _write_cell(ws, row, 1, "DEV Avr Rate TX:", font=LABEL_FONT)
            _write_cell(ws, row, 2, f"{test_r.dev_rate_tx:.2f} Mb/s", font=VALUE_FONT)
            row += 1
            _write_cell(ws, row, 1, "DEV Avr Rate RX:", font=LABEL_FONT)
            _write_cell(ws, row, 2, f"{test_r.dev_rate_rx:.2f} Mb/s", font=VALUE_FONT)
            row += 1

            if test_r.measured_throughput:
                tp_headers = ["Session", "Direction", "Avg (Mb/s)", "Min (Mb/s)", "Max (Mb/s)"]
                for ci, h in enumerate(tp_headers, start=1):
                    _write_cell(ws, row, ci, h, font=HEADER_FONT, fill=HEADER_FILL, border=THIN_BORDER, alignment=CENTER)
                row += 1
                for tp in test_r.measured_throughput:
                    _write_cell(ws, row, 1, tp.session_id, font=VALUE_FONT, border=THIN_BORDER, alignment=CENTER)
                    _write_cell(ws, row, 2, tp.throughput_direction, font=VALUE_FONT, border=THIN_BORDER, alignment=CENTER)
                    _write_cell(ws, row, 3, round(tp.throughput_avg, 2), font=VALUE_FONT, border=THIN_BORDER, alignment=CENTER)
                    _write_cell(ws, row, 4, round(tp.throughput_min, 2), font=VALUE_FONT, border=THIN_BORDER, alignment=CENTER)
                    _write_cell(ws, row, 5, round(tp.throughput_max, 2), font=VALUE_FONT, border=THIN_BORDER, alignment=CENTER)
                    row += 1
                row += 1

        if has_companion_tp:
            companion_dir = "RX" if test_r.direction == "TX" else "TX"
            _write_cell(ws, row, 1, f"Companion App Throughput ({companion_dir})", font=SUBHEADER_FONT, fill=SUBHEADER_FILL)
            row += 1
            _write_cell(ws, row, 1, "DEV Avr Rate TX:", font=LABEL_FONT)
            _write_cell(ws, row, 2, f"{test_r.companion_dev_rate_tx:.2f} Mb/s", font=VALUE_FONT)
            row += 1
            _write_cell(ws, row, 1, "DEV Avr Rate RX:", font=LABEL_FONT)
            _write_cell(ws, row, 2, f"{test_r.companion_dev_rate_rx:.2f} Mb/s", font=VALUE_FONT)
            row += 1

            if test_r.companion_throughput:
                tp_headers = ["Session", "Direction", "Avg (Mb/s)", "Min (Mb/s)", "Max (Mb/s)"]
                for ci, h in enumerate(tp_headers, start=1):
                    _write_cell(ws, row, ci, h, font=HEADER_FONT, fill=HEADER_FILL, border=THIN_BORDER, alignment=CENTER)
                row += 1
                for tp in test_r.companion_throughput:
                    _write_cell(ws, row, 1, tp.session_id, font=VALUE_FONT, border=THIN_BORDER, alignment=CENTER)
                    _write_cell(ws, row, 2, tp.throughput_direction, font=VALUE_FONT, border=THIN_BORDER, alignment=CENTER)
                    _write_cell(ws, row, 3, round(tp.throughput_avg, 2), font=VALUE_FONT, border=THIN_BORDER, alignment=CENTER)
                    _write_cell(ws, row, 4, round(tp.throughput_min, 2), font=VALUE_FONT, border=THIN_BORDER, alignment=CENTER)
                    _write_cell(ws, row, 5, round(tp.throughput_max, 2), font=VALUE_FONT, border=THIN_BORDER, alignment=CENTER)
                    row += 1
                row += 1

        # Separator
        row += 1

    # Adjust column widths
    for c in range(1, 15):
        ws.column_dimensions[get_column_letter(c)].width = 16
    ws.column_dimensions["A"].width = 22


# ──────────────────────────────────────────────────────────────────────
# Main
# ──────────────────────────────────────────────────────────────────────

def _find_log_files(path: str) -> List[str]:
    """Find all parseable log files from a path (file or directory).

    If path is a file, return it directly.
    If path is a directory, recursively search for all pytest.log files
    AND per-test .log files in subdirectories.  Per-test logs are created
    by the test framework alongside pytest.log and contain the same
    parseable markers.  When pytest.log is truncated (large runs), the
    per-test logs still hold the full results.
    Results are sorted by directory timestamp (oldest first) so that
    later entries overwrite earlier ones during deduplication.
    """
    if os.path.isfile(path):
        return [path]

    if not os.path.isdir(path):
        print(f"ERROR: Path not found: {path}")
        sys.exit(1)

    # Collect pytest.log files
    pytest_logs = sorted(glob.glob(os.path.join(path, "**/pytest.log"), recursive=True))

    # Collect per-test .log files (e.g. tests/dual/performance/test_...log)
    # These live in subdirectories within each timestamped run folder.
    all_logs = sorted(glob.glob(os.path.join(path, "**/*.log"), recursive=True))
    per_test_logs = [f for f in all_logs
                     if os.path.basename(f) != "pytest.log"
                     and "test_vf_multisession_fps_variants_dualhost" in f]

    log_files = sorted(set(pytest_logs + per_test_logs))
    # Filter out 'latest' symlink directory to avoid duplicate parsing
    log_files = [f for f in log_files if "/latest/" not in f]
    if not log_files:
        print(f"ERROR: No log files found in {path}")
        sys.exit(1)

    return log_files


def _get_dir_timestamp(log_path: str) -> str:
    """Extract the timestamp directory name from a log file path.

    Expected structure:
      .../logs/performance/2026-02-09T13:22:24/pytest.log
      .../logs/performance/2026-02-09T13:22:24/tests/dual/.../test_*.log
    Walks up parent directories to find the timestamp component.
    Returns the timestamp string for sorting, or empty string if not found.
    """
    parts = os.path.normpath(log_path).split(os.sep)
    ts_re = re.compile(r"\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}")
    for part in reversed(parts):
        if ts_re.match(part):
            return part
    return ""


def _dedup_results(results: List[TestResult]) -> List[TestResult]:
    """Keep only the latest result for each unique test case.

    Test cases are identified by their test_name which includes function,
    session count, FPS, and DSA variant (e.g. test_dualhost_vf_tx_fps_variants_single_core[8sess-59fps-dsa]).
    When duplicates exist, the last entry wins (caller must sort log files oldest-first).
    """
    seen: Dict[str, TestResult] = OrderedDict()
    for r in results:
        seen[r.test_name] = r
    deduped = list(seen.values())
    if len(deduped) < len(results):
        print(f"  Deduplicated: {len(results)} -> {len(deduped)} results (kept latest for each test case)")
    return deduped


def generate_report(input_path: str, output_path: Optional[str] = None,
                    platform_config_path: Optional[str] = None) -> str:
    """Generate the full XLSX report from pytest log file(s).

    Args:
        input_path: Path to a pytest.log file or a directory containing
                    pytest.log files in timestamped subdirectories.
        output_path: Optional output XLSX path. Auto-generated if None.
        platform_config_path: Optional path to platform_config.json.

    Returns:
        Path to the generated XLSX file.
    """
    log_files = _find_log_files(input_path)
    print(f"Found {len(log_files)} log file(s) to process")

    # Load platform configuration (look inside log dir first for auto-collected config)
    log_dir = input_path if os.path.isdir(input_path) else os.path.dirname(input_path)
    platform_config = _load_platform_config(platform_config_path, log_dir=log_dir)

    all_results: List[TestResult] = []
    for log_path in log_files:
        dir_ts = _get_dir_timestamp(log_path)
        basename = os.path.basename(log_path)
        if dir_ts:
            label = f"{dir_ts}/{basename}"
        else:
            label = log_path
        print(f"  Parsing {label} ...")
        results = parse_pytest_log(log_path)
        print(f"    {len(results)} test results")
        all_results.extend(results)

    if not all_results:
        print("WARNING: No test results found in any log file.")
        print("  Make sure the logs contain output from test_vf_multisession_fps_variants_dualhost.py")
        sys.exit(1)

    # Deduplicate - keep latest result for each test case
    results = _dedup_results(all_results)

    # Summary stats
    passed = sum(1 for r in results if r.overall_pass)
    failed = len(results) - passed
    print(f"  Total: {len(results)} unique test results (Passed: {passed}, Failed: {failed})")

    categories = categorize_results(results)
    for cat, cat_results in categories.items():
        if cat_results:
            print(f"  {cat}: {len(cat_results)} results")

    # Generate output path if not specified
    if output_path is None:
        if os.path.isdir(input_path):
            base_dir = input_path
        else:
            base_dir = os.path.dirname(input_path) or "."
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = os.path.join(base_dir, f"performance_report_{timestamp}.xlsx")

    # Create workbook
    wb = Workbook()

    write_summary_sheet(wb, results, categories, input_path, platform_config)

    for sheet_name in DATA_SHEET_NAMES:
        cat_results = categories.get(sheet_name, [])
        if cat_results:
            write_data_sheet(wb, sheet_name, cat_results)

    # Save
    wb.save(output_path)
    print(f"\nReport saved to: {output_path}")
    return output_path


def main():
    parser = argparse.ArgumentParser(
        description="Generate XLSX performance report from pytest log files.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument(
        "log_path",
        help="Path to a pytest.log file or directory containing log subdirectories",
    )
    parser.add_argument(
        "-o", "--output",
        help="Output XLSX file path (auto-generated if not specified)",
        default=None,
    )
    parser.add_argument(
        "-p", "--platform-config",
        help="Path to platform_config.json with SW/HW details (auto-detected if not specified)",
        default=None,
    )

    args = parser.parse_args()
    generate_report(args.log_path, args.output, args.platform_config)


if __name__ == "__main__":
    main()
