from __future__ import annotations

import csv
import html
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .common import ROOT, read_env


NOISE_PROFILE_PURPOSES = {
    "host-a": (
        "Host process outside Kubernetes: two NUMA-bound stress-ng STREAM groups, one per socket, "
        "about 5 busy cores. Runs outside kubepods so it can also take CPU time from pinned "
        "FFmpeg containers - the only profile RDT cannot fully repair."
    ),
    "pod-a": (
        "Two Guaranteed neighbor Pods, one per socket, 5 exclusive CPUs each, 32 MiB STREAM "
        "working set. Light DRAM pressure; interference is purely through shared LLC, memory "
        "controllers, and interconnect."
    ),
    "pod-b": (
        "Two Guaranteed neighbor Pods, one per socket, 12 exclusive CPUs each, 196 MiB STREAM "
        "working set. Heavy DRAM pressure with a working set well beyond the LLC."
    ),
    "pod-c": (
        "Two Guaranteed neighbor Pods, one per socket, 24 exclusive CPUs each, 512 MiB STREAM "
        "working set. Worst realistic co-tenant: saturates LLC and memory controllers together."
    ),
}


RDT_PROFILE_REFERENCE = [
    {
        "profile": "none",
        "mechanism": "Monitoring only (CMT/MBM)",
        "schemata": "Nothing is written. Groups are created under resctrl/mon_groups/, which allocates an RMID but no CLOS.",
        "effect": "Workload and noise share LLC and memory bandwidth exactly as the OS schedules them. Cache occupancy and bandwidth are only attributed per group.",
        "use": "Every clean reference run, and the unmanaged noisy run that every policy row is compared against.",
        "verify": "Non-zero occupancy and MBM for every group. A group reading zero for the whole window means its tasks were never associated, and the run is rejected.",
        "caveat": "Cannot improve or degrade performance. An FPS difference against another monitored run is workload variance, not RDT.",
    },
    {
        "profile": "cat-guarded",
        "mechanism": "L3 Cache Allocation Technology (CAT)",
        "schemata": "L3 mask fff0 for encoder and decoder, L3 mask f for noise, on every L3 domain. MB stays 100 everywhere. On a worker with 16 LLC ways that is 12 of 16 ways protected and 4 ways for the noise.",
        "effect": "Noise may only allocate lines inside its 4 ways, so it can no longer evict the FFmpeg working set from the other 12. The masks do not overlap, so the split is exclusive.",
        "use": "First cache lever to try. Works against cache thrashing and against streaming bandwidth noise that pollutes the LLC on its way to DRAM.",
        "verify": "rdt_noise_llc_occupancy_mib falls to roughly the noise mask size and stops growing while the workload occupancy holds. l3_cache_hit_ratio_pct should rise against the unmanaged row.",
        "caveat": "Caps footprint, not traffic: noise MBM barely moves, and CPU time taken by host-scope noise is not returned.",
    },
    {
        "profile": "cat-strong",
        "mechanism": "L3 Cache Allocation Technology (CAT)",
        "schemata": "L3 mask fffc for encoder and decoder, L3 mask 3 for noise. On a worker with 16 LLC ways that is 14 of 16 ways protected and 2 ways for the noise.",
        "effect": "Same mechanism as cat-guarded with a harsher split; the noise is left barely enough cache to run, which maximises protected capacity.",
        "use": "Only after cat-guarded has been measured and shown insufficient.",
        "verify": "Noise occupancy pinned near the 2-way limit. Compare workload occupancy and hit ratio against both the unmanaged row and the cat-guarded row.",
        "caveat": "Two ways is close to starvation, so the noise stops representing realistic contention. Always report this setting next to the measured noise occupancy.",
    },
    {
        "profile": "cat-16-1",
        "mechanism": "L3 Cache Allocation Technology (CAT), overlapping masks",
        "schemata": "L3 mask ffff (all 16 ways) for encoder and decoder, L3 mask 1 (one way, shared with the workload) for noise.",
        "effect": "The workload keeps the entire cache and the noise is squeezed into a single way it does not own exclusively. This is the strongest cache restriction the hardware allows without partitioning capacity away from FFmpeg.",
        "use": "Cases where cat-strong still leaves too much cache to the noise. On the reference platform this drove noise occupancy from 284 MiB to 4.8 MiB.",
        "verify": "Noise occupancy collapses to a few MiB. Workload occupancy should be unchanged or higher, because nothing was taken away from it.",
        "caveat": "Requires min_cbm_bits <= 1. The single way is shared, not reserved, so the noise still causes some replacement inside it.",
    },
    {
        "profile": "mba-80",
        "mechanism": "Memory Bandwidth Allocation (MBA)",
        "schemata": "MB 80 for the noise group, MB 100 for encoder and decoder, on every domain. L3 masks stay at the root default, so all classes still share the whole LLC.",
        "effect": "The hardware injects delay into memory requests issued by noise tasks, lowering the rate at which that class reaches DRAM. Throttling is per task, so it works even where noise shares CPUs with FFmpeg.",
        "use": "First bandwidth lever, and a light touch that keeps the noise realistic.",
        "verify": "rdt_noise_total_mbm_gbps drops against the unmanaged row while the workload MBM holds or rises. Unchanged noise MBM means the throttle did not bite.",
        "caveat": "The number is a hardware throttle control, not a guaranteed share of bandwidth. Nothing is reserved for FFmpeg; it only benefits from reduced competition.",
    },
    {
        "profile": "mba-60 / mba-40 / mba-10",
        "mechanism": "Memory Bandwidth Allocation (MBA)",
        "schemata": "MB 60, 40, or 10 for the noise group, MB 100 for encoder and decoder. LLC left shared.",
        "effect": "The same delay-injection mechanism at intermediate settings. mba-10 is the lowest value this hardware accepts (min_bandwidth=10, bandwidth_gran=10).",
        "use": "Sweep steps when mba-80 is too weak and mba-20 is stronger than necessary.",
        "verify": "Noise MBM should fall monotonically as the setting drops. If it does not, the limiter is not the noise memory rate.",
        "caveat": "The response is nonlinear: 60 does not mean 60% of bandwidth. Judge by measured MBM, not by the label.",
    },
    {
        "profile": "mba-20",
        "mechanism": "Memory Bandwidth Allocation (MBA)",
        "schemata": "MB 20 for the noise group, MB 100 for encoder and decoder. LLC left shared.",
        "effect": "Near-suppression of the noise memory request rate.",
        "use": "The single most effective policy measured on the reference platform. It alone restored 16 of 16 failing streams under the pod-b neighbor.",
        "verify": "Noise MBM close to the measurement floor while its CPU time continues, which separates bandwidth damage from CPU damage.",
        "caveat": "A strong intervention: at this setting the neighbor no longer represents a fully active co-tenant.",
    },
    {
        "profile": "<cat profile>+mba-<level>",
        "mechanism": "L3 CAT and MBA written into the same control group",
        "schemata": "Both lines set for one group, for example L3 ffff / MB 100 for encoder and decoder and L3 1 / MB 20 for noise (cat-16-1+mba-20).",
        "effect": "Bounds the noise on both dimensions at once: it can neither occupy protected cache ways nor issue memory requests at full rate.",
        "use": "Mixed noise that thrashes cache and consumes DRAM bandwidth at the same time. On the reference platform cat-16-1+mba-20 was the best policy for the host-a and pod-c neighbors.",
        "verify": "Both noise occupancy and noise MBM should fall together. Compare against the single-lever rows to see which one carried the recovery.",
        "caveat": "On its own it cannot attribute the recovery to cache or to bandwidth. Run the single-lever rows first when the goal is understanding rather than mitigation.",
    },
]


def rdt_profile_rows() -> list[dict[str, str]]:
    return list(RDT_PROFILE_REFERENCE)


def noise_profile_rows() -> list[dict[str, str]]:
    profiles = []
    for path in sorted((ROOT / "noisy-neighbors").glob("*.env")):
        cfg = read_env(path)
        profiles.append({
            "profile": path.stem,
            "scope": cfg.get("NOISY_NEIGHBOR_SCOPE", "pod"),
            "purpose": NOISE_PROFILE_PURPOSES.get(path.stem, "Custom stress-ng pressure profile."),
            "args": cfg.get("NOISY_NEIGHBOR_ARGS", ""),
            "cpu_request": cfg.get("NOISY_NEIGHBOR_CPU_REQUEST", ""),
            "memory_request": cfg.get("NOISY_NEIGHBOR_MEMORY_REQUEST", ""),
            "memory_limit": cfg.get("NOISY_NEIGHBOR_MEMORY_LIMIT", ""),
        })
    return profiles


def scan(root: Path, min_fps: float) -> list[dict[str, object]]:
    rows = []
    for config_path in root.rglob("config.json"):
        result = config_path.parent
        metrics_path = result / "metrics.csv"
        if not metrics_path.is_file():
            continue
        cfg = json.loads(config_path.read_text())
        fps = []
        ffmpeg_cpu_demand = {"encoder": [], "decoder": []}
        encoder_used_cpu_utilization = []
        decoder_used_cpu_utilization = []
        used_cpu_ids = {"encoder": set(), "decoder": set()}
        real_cpu_total_usage = {"encoder": [], "decoder": []}
        numa_metrics = {"cross_numa_upi_incoming_bytes_per_second": []}
        memory_metrics = {
            "l3_cache_hit_ratio": [],
            "l3_cache_misses_per_second": [],
            "dram_read_bytes_per_second": [],
            "dram_write_bytes_per_second": [],
        }
        with metrics_path.open(newline="") as handle:
            for row in csv.DictReader(handle):
                if (
                    row.get("category") == "FFmpeg progress"
                    and row["metric"] == "fps"
                    and row.get("role") == "encoder"
                ):
                    try:
                        parts = dict(item.split("=", 1) for item in row["value"].split(";"))
                        fps.append(float(parts["avg"]))
                    except (ValueError, KeyError):
                        pass
                if row["metric"] == "FFmpeg CPU demand":
                    try:
                        role = row.get("role", "")
                        if role in ffmpeg_cpu_demand:
                            ffmpeg_cpu_demand[role].append(float(row["value"]))
                    except ValueError:
                        pass
                if (
                    row["metric"] == "average FFmpeg utilization per measured used CPU"
                    and row.get("role") == "encoder"
                ):
                    try:
                        encoder_used_cpu_utilization.append(float(row["value"]))
                    except ValueError:
                        pass
                if row["metric"] in numa_metrics:
                    try:
                        parts = dict(item.split("=", 1) for item in row["value"].split(";"))
                        numa_metrics[row["metric"]].append(float(parts["avg"]))
                    except (ValueError, KeyError):
                        pass
                if row["metric"] in memory_metrics and row.get("scope", "") in ("", "{}"):
                    try:
                        parts = dict(item.split("=", 1) for item in row["value"].split(";"))
                        memory_metrics[row["metric"]].append(float(parts["avg"]))
                    except (ValueError, KeyError):
                        pass
                if (
                    row["metric"] == "average FFmpeg utilization per measured used CPU"
                    and row.get("role") == "decoder"
                ):
                    try:
                        decoder_used_cpu_utilization.append(float(row["value"]))
                    except ValueError:
                        pass
        cpu_usage_path = result / "cpu-usage.csv"
        if cpu_usage_path.is_file():
            with cpu_usage_path.open(newline="") as handle:
                for row in csv.DictReader(handle):
                    try:
                        if float(row["ffmpeg_cpu_seconds"]) > 0:
                            role = row.get("role", "")
                            if role in used_cpu_ids:
                                used_cpu_ids[role].add(int(row["cpu_id"]))
                    except (KeyError, ValueError):
                        pass
        system_usage_path = result / "core-system-usage.csv"
        if system_usage_path.is_file():
            with system_usage_path.open(newline="") as handle:
                for row in csv.DictReader(handle):
                    try:
                        cpu_id = int(row["cpu_id"])
                        # real_core_total_usage_pct is the pre-rename column name;
                        # a result directory collected before the rename still has it.
                        usage = float(row.get("real_cpu_total_usage_pct") or row["real_core_total_usage_pct"])
                        for role in real_cpu_total_usage:
                            if cpu_id in used_cpu_ids[role]:
                                real_cpu_total_usage[role].append(usage)
                    except (KeyError, ValueError):
                        pass
        placement = cfg.get("PLACEMENT", "")
        if placement == "exclusive":
            encoder_cores_per_session = cfg.get("ENC_CORES", "")
        elif placement == "numa-pool":
            encoder_cores_per_session = "socket pool"
        else:
            encoder_cores_per_session = "unrestricted"
        noise_enabled = False
        noise_numa_status = "n/a"
        noise_same_numa = "n/a"
        noise_precheck_status = "n/a"
        if cfg.get("NOISY_NEIGHBOR_ENABLED", cfg.get("NOISY_NEIGHBOUR_ENABLED")) == "1":
            for evidence_name in ("noisy-neighbor-after.json", "noisy-neighbour-after.json"):
                evidence_path = result / evidence_name
                if not evidence_path.is_file():
                    continue
                try:
                    evidence = json.loads(evidence_path.read_text())
                    noise_enabled = evidence.get("phase") == "Running"
                    numa = evidence.get("numa", {}) if isinstance(evidence.get("numa"), dict) else {}
                    noise_numa_status = str(numa.get("status", "n/a"))
                    same_socket = numa.get("same_socket")
                    noise_same_numa = "yes" if same_socket is True else "no" if same_socket is False else "n/a"
                    precheck = numa.get("precheck", {}) if isinstance(numa.get("precheck"), dict) else {}
                    if precheck:
                        noise_precheck_status = str(precheck.get("status", "n/a"))
                except (json.JSONDecodeError, OSError):
                    noise_enabled = False
                break
        enc_min = min(fps) if fps else 0.0
        upi_bandwidth = numa_metrics["cross_numa_upi_incoming_bytes_per_second"]
        upi_avg_gbps = sum(upi_bandwidth) / len(upi_bandwidth) / 1e9 if upi_bandwidth else None
        stream_count = int(cfg.get("STREAMS", 0))

        def memory_avg(metric: str) -> float | None:
            values = memory_metrics[metric]
            return sum(values) / len(values) if values else None

        hit_ratio = memory_avg("l3_cache_hit_ratio")
        l3_misses = memory_avg("l3_cache_misses_per_second")
        dram_read_gbps = memory_avg("dram_read_bytes_per_second")
        dram_write_gbps = memory_avg("dram_write_bytes_per_second")
        dram_total_gbps = (dram_read_gbps + dram_write_gbps) / 1e9 if dram_read_gbps is not None and dram_write_gbps is not None else None
        platform_spec = {}
        host_path = result / "host.json"
        if host_path.is_file():
            try:
                spec = json.loads(host_path.read_text()).get("platform_spec")
                platform_spec = spec if isinstance(spec, dict) else {}
            except (json.JSONDecodeError, OSError):
                platform_spec = {}
        theoretical_peak_gbps = platform_spec.get("theoretical_dram_gbps")
        dram_peak_pct = (
            dram_total_gbps / theoretical_peak_gbps * 100
            if dram_total_gbps is not None and theoretical_peak_gbps else None
        )
        rdt_group_llc: dict[str, list[float]] = {"encoder": [], "decoder": [], "workload": [], "noise": []}
        rdt_group_mbm: dict[str, list[float]] = {"encoder": [], "decoder": [], "workload": [], "noise": []}
        rdt_path = result / "rdt-summary.json"
        if rdt_path.is_file():
            try:
                for item in json.loads(rdt_path.read_text()):
                    group = item.get("group")
                    if group in rdt_group_llc:
                        llc = float(item.get("llc_occupancy_bytes_avg", 0))
                        mbm = float(item.get("mbm_total_bytes_per_second_avg", 0))
                        rdt_group_llc[group].append(llc)
                        rdt_group_mbm[group].append(mbm)
                        if group in ("encoder", "decoder"):
                            # Keep the aggregate workload columns meaningful for per-role runs.
                            rdt_group_llc["workload"].append(llc)
                            rdt_group_mbm["workload"].append(mbm)
            except (json.JSONDecodeError, OSError, TypeError, ValueError):
                pass

        def rdt_llc_mib(group: str) -> object:
            values = rdt_group_llc[group]
            return round(sum(values) / 2**20, 3) if values else "unavailable"

        def rdt_mbm_gbps(group: str) -> object:
            values = rdt_group_mbm[group]
            return round(sum(values) / 1e9, 3) if values else "unavailable"
        rows.append({
            "result_directory": str(result), "scenario": cfg.get("SCENARIO", ""), "scenario_description": cfg.get("DESCRIPTION", ""),
            "cpu_placement": cfg.get("PLACEMENT", ""), "worker_node": cfg.get("NODE", cfg.get("LAB_DEFAULT_NODE", "")),
            "streams": stream_count, "resolution": cfg.get("RESOLUTION", ""), "preset": cfg.get("PRESET", ""),
            "noisy_neighbor": "enabled" if noise_enabled else "",
            "noisy_neighbor_profile": cfg.get("NOISY_NEIGHBOR_PROFILE", cfg.get("NOISY_NEIGHBOUR_PROFILE", "")) if noise_enabled else "",
            "noisy_neighbor_image": cfg.get("NOISY_NEIGHBOR_IMAGE", cfg.get("NOISY_NEIGHBOUR_IMAGE", "")) if noise_enabled else "",
            "noisy_neighbor_args": cfg.get("NOISY_NEIGHBOR_ARGS", cfg.get("NOISY_NEIGHBOUR_ARGS", "")) if noise_enabled else "",
            "noisy_neighbor_numa_status": noise_numa_status if noise_enabled else "n/a",
            "noisy_neighbor_same_numa": noise_same_numa if noise_enabled else "n/a",
            "noisy_neighbor_precheck_status": noise_precheck_status if noise_enabled else "n/a",
            "rdt_monitor": "enabled" if cfg.get("RDT_MONITOR") == "1" or cfg.get("RDT_CONTROL_PROFILE", "none") != "none" else "",
            "rdt_control_profile": cfg.get("RDT_CONTROL_PROFILE", "none"),
            "rdt_workload_llc_occupancy_mib": rdt_llc_mib("workload"),
            "rdt_workload_total_mbm_gbps": rdt_mbm_gbps("workload"),
            "rdt_encoder_llc_occupancy_mib": rdt_llc_mib("encoder"),
            "rdt_encoder_total_mbm_gbps": rdt_mbm_gbps("encoder"),
            "rdt_decoder_llc_occupancy_mib": rdt_llc_mib("decoder"),
            "rdt_decoder_total_mbm_gbps": rdt_mbm_gbps("decoder"),
            "rdt_noise_llc_occupancy_mib": rdt_llc_mib("noise"),
            "rdt_noise_total_mbm_gbps": rdt_mbm_gbps("noise"),
            "rdt_focus_session": cfg.get("RDT_FOCUS_SESSION", ""),
            "encoder_cores_per_session": encoder_cores_per_session, "encoder_threads": cfg.get("ENC_THREADS", "auto"),
            "slices": cfg.get("SLICES", "auto"), "sliced_threads": cfg.get("SLICED_THREADS", "default"),
            "x264_overrides": cfg.get("X264_EXTRA", ""), "bitrate": cfg.get("BITRATE", ""),
            "minimum_encoder_fps": round(enc_min, 3),
            "encoder_cpu_busy_core_equivalents": round(sum(ffmpeg_cpu_demand["encoder"]), 3) if ffmpeg_cpu_demand["encoder"] else "",
            "encoder_real_cpu_total_usage_avg_pct": round(sum(real_cpu_total_usage["encoder"]) / len(real_cpu_total_usage["encoder"]), 3) if real_cpu_total_usage["encoder"] else "",
            "encoder_avg_utilization_per_used_cpu_pct": round(sum(encoder_used_cpu_utilization) / len(encoder_used_cpu_utilization), 3) if encoder_used_cpu_utilization else "",
            "decoder_cpu_busy_core_equivalents": round(sum(ffmpeg_cpu_demand["decoder"]), 3) if ffmpeg_cpu_demand["decoder"] else "",
            "decoder_real_cpu_total_usage_avg_pct": round(sum(real_cpu_total_usage["decoder"]) / len(real_cpu_total_usage["decoder"]), 3) if real_cpu_total_usage["decoder"] else "",
            "decoder_avg_utilization_per_used_cpu_pct": round(sum(decoder_used_cpu_utilization) / len(decoder_used_cpu_utilization), 3) if decoder_used_cpu_utilization else "",
            "cross_numa_upi_bandwidth_gbps": round(upi_avg_gbps, 3) if upi_avg_gbps is not None else "unavailable",
            "cross_numa_upi_gbps_per_stream": round(upi_avg_gbps / stream_count, 3) if upi_avg_gbps is not None and stream_count else "unavailable",
            "l3_cache_hit_ratio_pct": round(hit_ratio * 100, 3) if hit_ratio is not None else "unavailable",
            "l3_cache_misses_per_second_millions": round(l3_misses / 1e6, 3) if l3_misses is not None else "unavailable",
            "dram_read_gbps": round(dram_read_gbps / 1e9, 3) if dram_read_gbps is not None else "unavailable",
            "dram_write_gbps": round(dram_write_gbps / 1e9, 3) if dram_write_gbps is not None else "unavailable",
            "dram_total_gbps": round(dram_total_gbps, 3) if dram_total_gbps is not None else "unavailable",
            "dram_total_gbps_per_stream": round(dram_total_gbps / stream_count, 3) if dram_total_gbps is not None and stream_count else "unavailable",
            "theoretical_dram_peak_gbps": theoretical_peak_gbps if theoretical_peak_gbps else "unavailable",
            "dram_pct_of_theoretical_peak": round(dram_peak_pct, 2) if dram_peak_pct is not None else "unavailable",
            "platform_spec": platform_spec,
            "pass": enc_min >= min_fps,
        })
    return sorted(rows, key=lambda row: (not bool(row["pass"]), -int(row["streams"]), -float(row["minimum_encoder_fps"])))


def build_summary(root: Path, min_fps: float) -> tuple[Path, Path]:
    rows = scan(root, min_fps)
    standard_rows = [row for row in rows if not row["noisy_neighbor"]]
    noise_rows = [row for row in rows if row["noisy_neighbor"]]

    def cpu_demand(row: dict[str, object]) -> float:
        try:
            return float(row["encoder_cpu_busy_core_equivalents"]) + float(row["decoder_cpu_busy_core_equivalents"])
        except (TypeError, ValueError):
            return float("inf")

    def winner_key(row: dict[str, object]) -> tuple[float, float, float]:
        return (-int(row["streams"]), cpu_demand(row), -float(row["minimum_encoder_fps"]))
    csv_path = root / "summary.csv"
    fields = [
        "scenario", "cpu_placement", "streams", "resolution", "preset", "bitrate",
        "noisy_neighbor", "noisy_neighbor_profile",
        "noisy_neighbor_numa_status", "noisy_neighbor_same_numa", "noisy_neighbor_precheck_status",
        "rdt_monitor", "rdt_control_profile", "rdt_focus_session",
        "rdt_workload_llc_occupancy_mib", "rdt_workload_total_mbm_gbps",
        "rdt_encoder_llc_occupancy_mib", "rdt_encoder_total_mbm_gbps",
        "rdt_decoder_llc_occupancy_mib", "rdt_decoder_total_mbm_gbps",
        "rdt_noise_llc_occupancy_mib", "rdt_noise_total_mbm_gbps",
        "encoder_cores_per_session", "encoder_threads", "slices", "minimum_encoder_fps",
        "encoder_cpu_busy_core_equivalents", "encoder_real_cpu_total_usage_avg_pct",
        "encoder_avg_utilization_per_used_cpu_pct",
        "decoder_cpu_busy_core_equivalents", "decoder_real_cpu_total_usage_avg_pct",
        "decoder_avg_utilization_per_used_cpu_pct", "cross_numa_upi_bandwidth_gbps",
        "cross_numa_upi_gbps_per_stream",
        "l3_cache_hit_ratio_pct", "l3_cache_misses_per_second_millions",
        "dram_read_gbps", "dram_write_gbps", "dram_total_gbps", "dram_total_gbps_per_stream",
        "theoretical_dram_peak_gbps", "dram_pct_of_theoretical_peak",
        "pass",
        "worker_node", "sliced_threads", "x264_overrides",
        "noisy_neighbor_image", "noisy_neighbor_args", "result_directory",
    ]
    display_names = {
        "scenario": "Scenario",
        "cpu_placement": "CPU placement",
        "streams": "Streams",
        "resolution": "Resolution",
        "preset": "Preset",
        "bitrate": "Bitrate",
        "noisy_neighbor": "Noisy neighbor",
        "noisy_neighbor_profile": "Noise profile",
        "noisy_neighbor_numa_status": "NN socket placement",
        "noisy_neighbor_same_numa": "NN same NUMA socket",
        "noisy_neighbor_precheck_status": "NN precheck socket placement",
        "noisy_neighbor_image": "stress-ng image",
        "noisy_neighbor_args": "stress-ng arguments",
        "rdt_monitor": "RDT monitor",
        "rdt_control_profile": "RDT control profile",
        "rdt_workload_llc_occupancy_mib": "RDT workload LLC occupancy (MiB)",
        "rdt_workload_total_mbm_gbps": "RDT workload-group MBM total (GB/s)",
        "rdt_focus_session": "RDT focus session",
        "rdt_encoder_llc_occupancy_mib": "RDT encoder LLC occupancy (MiB)",
        "rdt_encoder_total_mbm_gbps": "RDT encoder-group MBM total (GB/s)",
        "rdt_decoder_llc_occupancy_mib": "RDT decoder LLC occupancy (MiB)",
        "rdt_decoder_total_mbm_gbps": "RDT decoder-group MBM total (GB/s)",
        "rdt_noise_llc_occupancy_mib": "RDT noise LLC occupancy (MiB)",
        "rdt_noise_total_mbm_gbps": "RDT noise total MBM (GB/s)",
        "encoder_cores_per_session": "Encoder cores\nper session",
        "encoder_threads": "Encoder threads",
        "slices": "Slices",
        "minimum_encoder_fps": "Minimum encoder FPS",
        "encoder_cpu_busy_core_equivalents": "Encoder CPU busy\n(core-equivalents)",
        "encoder_real_cpu_total_usage_avg_pct": "Encoder-CPU real total usage\n(avg %)",
        "encoder_avg_utilization_per_used_cpu_pct": "Encoder avg utilization\nper used CPU (%)",
        "decoder_cpu_busy_core_equivalents": "Decoder CPU busy\n(core-equivalents)",
        "decoder_real_cpu_total_usage_avg_pct": "Decoder-CPU real total usage\n(avg %)",
        "decoder_avg_utilization_per_used_cpu_pct": "Decoder avg utilization\nper used CPU (%)",
        "cross_numa_upi_bandwidth_gbps": "Cross-socket UPI incoming bandwidth\n(GB/s)",
        "cross_numa_upi_gbps_per_stream": "Cross-socket UPI incoming\n(GB/s per stream)",
        "l3_cache_hit_ratio_pct": "Whole worker L3 cache\nhit ratio (%)",
        "l3_cache_misses_per_second_millions": "Whole worker L3 misses\n(M/s)",
        "dram_read_gbps": "Whole worker DRAM read\n(GB/s)",
        "dram_write_gbps": "Whole worker DRAM write\n(GB/s)",
        "dram_total_gbps": "Whole worker DRAM read+write\n(GB/s)",
        "dram_total_gbps_per_stream": "Whole worker DRAM read+write\n(GB/s per stream)",
        "theoretical_dram_peak_gbps": "Theoretical DRAM peak\n(GB/s)",
        "dram_pct_of_theoretical_peak": "Measured DRAM share\nof theoretical peak (%)",
        "pass": "Pass",
        "worker_node": "Worker node",
        "sliced_threads": "Sliced threads",
        "x264_overrides": "x264 overrides",
        "result_directory": "Result directory",
    }
    with csv_path.open("w", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        writer.writerows({field: row[field] for field in fields} for row in rows)
    workbook = Workbook(); sheet = workbook.active; sheet.title = "All results"; sheet.append([display_names[field] for field in fields])
    for row in rows: sheet.append([row[field] for field in fields])
    standard_sheet = workbook.create_sheet("No noise")
    standard_sheet.append([display_names[field] for field in fields])
    for row in rows:
        if not row["noisy_neighbor"]:
            standard_sheet.append([row[field] for field in fields])
    noise_sheet = workbook.create_sheet("Noisy neighbor")
    noise_sheet.append([display_names[field] for field in fields])
    for row in rows:
        if row["noisy_neighbor"]:
            noise_sheet.append([row[field] for field in fields])
    best = workbook.create_sheet("Best passing")
    best.append([display_names[field] for field in fields])
    by_key = {}
    for row in standard_rows:
        if row["pass"]:
            by_key.setdefault((row["scenario"], row["cpu_placement"], row["preset"]), row)
    for row in by_key.values(): best.append([row[field] for field in fields])
    scenario_winners = {}
    for row in standard_rows:
        if not row["pass"]:
            continue
        name = str(row["scenario"])
        if name not in scenario_winners or winner_key(row) < winner_key(scenario_winners[name]):
            scenario_winners[name] = row
    winners_sheet = workbook.create_sheet("Best by scenario")
    winner_fields = [
        "scenario", "cpu_placement", "streams", "minimum_encoder_fps", "preset",
        "noisy_neighbor", "noisy_neighbor_profile", "noisy_neighbor_numa_status",
        "encoder_threads", "slices", "encoder_cpu_busy_core_equivalents",
        "encoder_real_cpu_total_usage_avg_pct", "decoder_real_cpu_total_usage_avg_pct",
        "cross_numa_upi_bandwidth_gbps", "cross_numa_upi_gbps_per_stream",
        "result_directory",
    ]
    winners_sheet.append([display_names[field] for field in winner_fields])
    for row in scenario_winners.values():
        winners_sheet.append([row[field] for field in winner_fields])
    autosize_widths = {
        "scenario": 20, "cpu_placement": 16, "streams": 10, "minimum_encoder_fps": 18,
        "preset": 12, "encoder_threads": 16, "slices": 10,
        "noisy_neighbor": 16, "noisy_neighbor_profile": 18, "noisy_neighbor_numa_status": 22,
        "encoder_cpu_busy_core_equivalents": 22,
        "encoder_real_cpu_total_usage_avg_pct": 22,
        "decoder_real_cpu_total_usage_avg_pct": 22,
        "cross_numa_upi_bandwidth_gbps": 22, "cross_numa_upi_gbps_per_stream": 22,
        "result_directory": 55,
    }
    winners_sheet.freeze_panes = "A2"
    winners_sheet.auto_filter.ref = winners_sheet.dimensions
    winners_sheet.sheet_view.showGridLines = False
    winners_sheet.row_dimensions[1].height = 34
    for cell in winners_sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for column, field in enumerate(winner_fields, 1):
        winners_sheet.column_dimensions[get_column_letter(column)].width = autosize_widths[field]
    scenario_descriptions = {}
    for row in rows:
        scenario_descriptions.setdefault(str(row["scenario"]), str(row["scenario_description"]))
    scenario_details = {
        "baseline": {
            "placement": "Free scheduling. No taskset, no NUMA hint; threads may migrate across all container-allowed CPUs and both sockets.",
            "codec": "x264 medium preset, 15 encoder threads, 2 sliced-threaded output slices, automatic decoder and filter threads.",
            "cpu": "Burstable Pods: 500m CPU request per decoder and encoder, no CPU limit, no exclusive CPU ownership.",
            "purpose": "Unoptimized reference. Every optimization is measured as a delta against this row.",
            "tradeoff": "Lowest determinism and lowest density; cross-socket migration inflates interconnect traffic.",
        },
        "numa-pool": {
            "placement": "Taskset to all usable CPUs of one socket. Sessions alternate sockets, so each session stays NUMA-local but shares its socket pool.",
            "codec": "Identical to baseline: x264 medium, 15 encoder threads, 2 sliced-threaded slices.",
            "cpu": "Burstable Pods: 500m CPU request per decoder and encoder. No per-stream exclusive cores.",
            "purpose": "Isolates the value of NUMA locality alone, with the codec settings and QoS class held constant against baseline.",
            "tradeoff": "Removes cross-socket traffic and keeps elasticity inside the socket, but cannot borrow idle CPU from the other socket.",
        },
        "pinned": {
            "placement": "Guaranteed Pods. The kubelet static CPU Manager assigns exclusive cores and the single-numa-node Topology Manager keeps each container NUMA-local. No taskset: kubelet owns the cpuset.",
            "codec": "x264 medium, 1 decoder thread, 1 filter thread, 15 encoder frame threads (SLICED_THREADS=0), 2 output slices.",
            "cpu": "1 exclusive decoder CPU plus 5 exclusive encoder CPUs per stream; integer requests equal limits on every container.",
            "purpose": "Full CPU QoS. Measures deterministic exclusive placement plus NUMA locality.",
            "tradeoff": "Highest density and repeatability, but idle assigned capacity cannot help other streams and worker topology caps the stream count.",
        },
    }
    legend_sheet = workbook.create_sheet("Scenario legend")
    legend_sheet.append(["scenario", "description"])
    for name, description in sorted(scenario_descriptions.items()):
        legend_sheet.append([name, description])
    for target in (sheet, standard_sheet, noise_sheet, best):
        target.freeze_panes = "A2"
        target.auto_filter.ref = target.dimensions
        target.sheet_view.showGridLines = False
        for cell in target[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
        for row_number in range(2, target.max_row + 1):
            passed = bool(target.cell(row_number, fields.index("pass") + 1).value)
            fill = PatternFill("solid", fgColor="E7F6EA" if passed else "FDEBEC")
            for cell in target[row_number]:
                cell.fill = fill
        target.row_dimensions[1].height = 34
        for cell in target[1]:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        widths = {
            "scenario": 18, "cpu_placement": 15, "streams": 10, "resolution": 12,
            "preset": 12, "bitrate": 10, "encoder_cores_per_session": 14,
            "noisy_neighbor": 16, "noisy_neighbor_profile": 18,
            "noisy_neighbor_numa_status": 22, "noisy_neighbor_same_numa": 20, "noisy_neighbor_precheck_status": 24,
            "rdt_monitor": 12, "rdt_control_profile": 18,
            "rdt_workload_llc_occupancy_mib": 22,
            "rdt_workload_total_mbm_gbps": 22,
            "rdt_focus_session": 16,
            "rdt_encoder_llc_occupancy_mib": 22,
            "rdt_encoder_total_mbm_gbps": 22,
            "rdt_decoder_llc_occupancy_mib": 22,
            "rdt_decoder_total_mbm_gbps": 22,
            "rdt_noise_llc_occupancy_mib": 22,
            "rdt_noise_total_mbm_gbps": 22,
            "noisy_neighbor_image": 36, "noisy_neighbor_args": 55,
            "encoder_threads": 15, "slices": 9, "minimum_encoder_fps": 16,
            "encoder_cpu_busy_core_equivalents": 20,
            "encoder_real_cpu_total_usage_avg_pct": 20,
            "encoder_avg_utilization_per_used_cpu_pct": 20,
            "decoder_cpu_busy_core_equivalents": 20,
            "decoder_real_cpu_total_usage_avg_pct": 20,
            "decoder_avg_utilization_per_used_cpu_pct": 20, "pass": 9,
            "cross_numa_upi_bandwidth_gbps": 22,
            "cross_numa_upi_gbps_per_stream": 22,
            "l3_cache_hit_ratio_pct": 20,
            "l3_cache_misses_per_second_millions": 20,
            "dram_read_gbps": 20, "dram_write_gbps": 20,
            "dram_total_gbps": 22, "dram_total_gbps_per_stream": 22,
            "theoretical_dram_peak_gbps": 20, "dram_pct_of_theoretical_peak": 22,
            "worker_node": 15, "sliced_threads": 14, "x264_overrides": 22,
            "result_directory": 55,
        }
        for column, field in enumerate(fields, 1):
            target.column_dimensions[get_column_letter(column)].width = widths[field]
    legend_sheet.freeze_panes = "A2"
    legend_sheet.sheet_view.showGridLines = False
    for cell in legend_sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    legend_sheet.column_dimensions["A"].width = 24
    legend_sheet.column_dimensions["B"].width = 90
    platform_fields = [
        ("worker_node", "Worker node"),
        ("cpu_model", "CPU model"),
        ("sockets", "Sockets"),
        ("cores_per_socket", "Cores per socket"),
        ("threads_per_core", "Threads per core"),
        ("logical_cpus", "Logical CPUs"),
        ("numa_nodes", "NUMA nodes"),
        ("cpu_max_mhz", "CPU max MHz"),
        ("l3_cache", "L3 cache"),
        ("memory_total_gib", "Installed memory (GiB)"),
        ("populated_dimms", "Populated DIMMs"),
        ("dimm_size_gib", "DIMM size (GiB)"),
        ("dimm_type", "DIMM type"),
        ("populated_channels", "Populated channels"),
        ("channels_per_socket", "Channels per socket"),
        ("memory_transfer_mt_s", "Memory transfer (MT/s)"),
        ("memory_transfer_source", "Transfer rate source"),
        ("theoretical_dram_gbps", "Theoretical DRAM peak (GB/s)"),
    ]
    platform_by_node: dict[str, dict[str, object]] = {}
    for row in rows:
        spec = row.get("platform_spec")
        if isinstance(spec, dict) and spec:
            platform_by_node.setdefault(str(row["worker_node"]), dict(spec, worker_node=row["worker_node"]))
    platform_sheet = workbook.create_sheet("Worker platform")
    platform_sheet.append([label for _, label in platform_fields])
    for spec in platform_by_node.values():
        platform_sheet.append([spec.get(field, "unavailable") for field, _ in platform_fields])
    platform_sheet.freeze_panes = "A2"
    platform_sheet.sheet_view.showGridLines = False
    for cell in platform_sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for column in range(1, len(platform_fields) + 1):
        platform_sheet.column_dimensions[get_column_letter(column)].width = 22
    profile_sheet = workbook.create_sheet("Noise profile legend")
    profile_fields = ["profile", "scope", "purpose", "args", "cpu_request", "memory_request", "memory_limit"]
    profile_sheet.append(["Profile", "Scope", "Purpose", "stress-ng arguments", "CPU request", "Memory request", "Memory limit"])
    for profile in noise_profile_rows():
        profile_sheet.append([profile[field] for field in profile_fields])
    profile_sheet.freeze_panes = "A2"
    profile_sheet.sheet_view.showGridLines = False
    for cell in profile_sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    for column, width in enumerate((20, 12, 72, 72, 16, 18, 16), 1):
        profile_sheet.column_dimensions[get_column_letter(column)].width = width
    rdt_sheet = workbook.create_sheet("RDT profile legend")
    rdt_fields = ["profile", "mechanism", "schemata", "effect", "use", "verify", "caveat"]
    rdt_sheet.append(["Profile", "RDT mechanism", "Applied schemata", "What it does", "When to use", "How to verify", "Limits and caveat"])
    for profile in rdt_profile_rows():
        rdt_sheet.append([profile[field] for field in rdt_fields])
    rdt_sheet.freeze_panes = "A2"
    rdt_sheet.sheet_view.showGridLines = False
    for cell in rdt_sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    for column, width in enumerate((16, 30, 70, 70, 46, 62, 70), 1):
        rdt_sheet.column_dimensions[get_column_letter(column)].width = width
    xlsx_path = root / "summary.xlsx"; workbook.save(xlsx_path)
    html_path = root / "summary.html"

    def render_cell(row: dict[str, object], field: str) -> str:
        value = row[field]
        if field == "pass":
            label = "PASS" if value else "FAIL"
            return f"<td><span class='status {'ok' if value else 'bad'}'><i></i>{label}</span></td>"
        if field == "result_directory":
            try:
                relative = Path(str(value)).relative_to(root) / "report.html"
                href = html.escape(relative.as_posix(), quote=True)
                return f"<td><a class='report-link' href='{href}'>View report <span>↗</span></a></td>"
            except ValueError:
                pass
        suffix = "%" if field in (
            "encoder_real_cpu_total_usage_avg_pct",
            "encoder_avg_utilization_per_used_cpu_pct",
            "decoder_real_cpu_total_usage_avg_pct",
            "decoder_avg_utilization_per_used_cpu_pct",
            "l3_cache_hit_ratio_pct",
            "dram_pct_of_theoretical_peak",
        ) and value not in ("", "unavailable") else ""
        css = "numeric" if field in (
            "streams", "encoder_threads", "slices",
            "minimum_encoder_fps", "encoder_cpu_busy_core_equivalents",
            "encoder_real_cpu_total_usage_avg_pct", "encoder_avg_utilization_per_used_cpu_pct",
            "decoder_cpu_busy_core_equivalents",
            "decoder_real_cpu_total_usage_avg_pct",
            "decoder_avg_utilization_per_used_cpu_pct",
            "cross_numa_upi_bandwidth_gbps", "cross_numa_upi_gbps_per_stream",
            "l3_cache_hit_ratio_pct", "l3_cache_misses_per_second_millions",
            "dram_read_gbps", "dram_write_gbps", "dram_total_gbps", "dram_total_gbps_per_stream",
            "theoretical_dram_peak_gbps", "dram_pct_of_theoretical_peak",
        ) else ""
        return f"<td class='{css}'>{html.escape(str(value))}{suffix}</td>"

    def result_rows_html(result_rows: list[dict[str, object]]) -> str:
        return "".join(
            f"<tr class='{'pass' if row['pass'] else 'fail'}'>"
            + "".join(render_cell(row, field) for field in fields) + "</tr>"
            for row in result_rows
        ) or f"<tr><td colspan='{len(fields)}'>No results.</td></tr>"

    standard_trs = result_rows_html(standard_rows)
    noise_trs = result_rows_html(noise_rows)
    passing = [row for row in standard_rows if row["pass"]]
    top = min(passing, key=winner_key) if passing else None
    best_streams = str(top["streams"]) if top else "none"
    best_scenario = f"{top['scenario']} / {top['cpu_placement']}" if top else "none"

    def numa_value(row: dict[str, object], field: str, suffix: str) -> str:
        value = row[field]
        return "unavailable" if value in ("", "unavailable") else f"{value}{suffix}"

    def real_cpu_usage(row: dict[str, object], role: str) -> str:
        value = row[f"{role}_real_cpu_total_usage_avg_pct"]
        return f"{float(value):.3f}%" if value != "" else "unavailable"

    winner_cards = "".join(
        f"<article class='winner {'recommended' if row is top else ''}'>"
        + ("<div class='winner-badge'>Best overall</div>" if row is top else "")
        + f"<div class='winner-name'>{html.escape(name)}</div>"
        + f"<div class='winner-score'><strong>{row['streams']}</strong><span>passing streams</span></div>"
        + "<dl>"
        + f"<div><dt>Minimum FPS</dt><dd>{float(row['minimum_encoder_fps']):.3f}</dd></div>"
        + f"<div><dt>Placement</dt><dd>{html.escape(str(row['cpu_placement']))}</dd></div>"
        + f"<div><dt>Preset</dt><dd>{html.escape(str(row['preset']))}</dd></div>"
        + f"<div><dt>Threads / slices</dt><dd>{html.escape(str(row['encoder_threads']))} / {html.escape(str(row['slices']))}</dd></div>"
        + f"<div><dt>Encoder-CPU real usage</dt><dd>{real_cpu_usage(row, 'encoder')}</dd></div>"
        + f"<div><dt>Decoder-CPU real usage</dt><dd>{real_cpu_usage(row, 'decoder')}</dd></div>"
        + f"<div><dt>Cross-socket UPI incoming</dt><dd>{numa_value(row, 'cross_numa_upi_bandwidth_gbps', ' GB/s')}</dd></div>"
        + f"<div><dt>UPI incoming per stream</dt><dd>{numa_value(row, 'cross_numa_upi_gbps_per_stream', ' GB/s')}</dd></div>"
        + f"<div><dt>Encoder CPU</dt><dd>{row['encoder_cpu_busy_core_equivalents']} cores</dd></div>"
        + f"<div><dt>Total FFmpeg CPU</dt><dd>{cpu_demand(row):.3f} cores</dd></div>"
        + "</dl>"
        + render_cell(row, "result_directory").replace("<td>", "<div class='winner-link'>").replace("</td>", "</div>")
        + "</article>"
        for name, row in scenario_winners.items()
    ) or "<p>No passing scenarios yet.</p>"
    scenario_rows = "".join(
        "<tr>"
        f"<td><strong>{html.escape(name)}</strong><small>{html.escape(description or 'No description')}</small></td>"
        + "".join(f"<td>{html.escape(details[key])}</td>" for key in ("placement", "codec", "cpu", "purpose", "tradeoff"))
        + "</tr>"
        for name, description in sorted(scenario_descriptions.items())
        for details in [scenario_details.get(name, {
            "placement": "See captured configuration.", "codec": "See captured FFmpeg command.",
            "cpu": "See CPU placement evidence.", "purpose": description or "Custom scenario.",
            "tradeoff": "Not documented.",
        })]
    )
    metric_legend = "".join([
        "<div class='legend-item'><strong>encoder_cores_per_session</strong><span>Physical-core allowance for one encoder process, from ENC_CORES. Exclusive placement shows the configured pinned core count, such as 5, 6, or 7; the CPU request the kubelet sees is that count times LAB_THREADS_PER_CORE, so with SMT on it is twice this number of logical CPUs. NUMA-pool shows socket pool; free scheduling shows unrestricted. Campaign-wide CPU union is intentionally not shown.</span></div>",
        "<div class='legend-item'><strong>encoder_cpu_busy_core_equivalents</strong><span>Sum of measured encoder FFmpeg process CPU time divided by measurement duration. Excludes decoder, Kubernetes, kernel, and unrelated workloads.</span></div>",
        "<div class='legend-item'><strong>encoder_real_cpu_total_usage_avg_pct</strong><span>Arithmetic mean of exact-window /proc/stat busy percentages across encoder-used logical CPUs only. Includes FFmpeg, Kubernetes, kernel, exporters, and all other host work on those CPUs. The unit is a logical CPU: with LAB_THREADS_PER_CORE=1 that is a physical core, and with SMT on each core contributes two CPUs to this average.</span></div>",
        "<div class='legend-item'><strong>encoder_avg_utilization_per_used_cpu_pct</strong><span>Encoder FFmpeg CPU demand divided by number of distinct encoder-used CPU IDs. Unit: percent.</span></div>",
        "<div class='legend-item'><strong>decoder_cpu_busy_core_equivalents</strong><span>Sum of measured decoder FFmpeg process CPU time divided by measurement duration. Excludes encoder, Kubernetes, kernel, and unrelated workloads.</span></div>",
        "<div class='legend-item'><strong>decoder_real_cpu_total_usage_avg_pct</strong><span>Arithmetic mean of exact-window /proc/stat busy percentages across decoder-used logical CPUs only. Includes FFmpeg, Kubernetes, kernel, exporters, and all other host work on those CPUs. The unit is a logical CPU: with LAB_THREADS_PER_CORE=1 that is a physical core, and with SMT on each core contributes two CPUs to this average.</span></div>",
        "<div class='legend-item'><strong>decoder_avg_utilization_per_used_cpu_pct</strong><span>Decoder FFmpeg CPU demand divided by number of distinct decoder-used CPU IDs. Unit: percent.</span></div>",
        "<div class='legend-item'><strong>cross_numa_upi_bandwidth_gbps</strong><span>Direct host-wide PCM incoming traffic summed across available UPI links and sockets, without adding outgoing copies. Best cross-socket traffic metric available; not process-attributed.</span></div>",
        "<div class='legend-item'><strong>cross_numa_upi_gbps_per_stream</strong><span>Direct UPI incoming bandwidth divided by benchmark stream count. Useful normalized comparison when same worker runs no unrelated traffic. This is traffic intensity, not percent remote-memory accesses.</span></div>",
        "<div class='legend-item'><strong>rdt_monitor</strong><span>Run-scoped CMT/MBM measurement groups were active. Monitoring alone changes no allocation and cannot affect the score.</span></div>",
        "<div class='legend-item'><strong>rdt_control_profile</strong><span>Applied isolation policy. <code>none</code> means measurement only. <code>cat-*</code> partitions LLC ways between workload and noise. <code>mba-N</code> throttles the noise memory-bandwidth class to control value N. Full meaning of each value is in the RDT control profile reference table below.</span></div>",
        "<div class='legend-item'><strong>rdt_workload_llc_occupancy_mib</strong><span>Cache footprint held by FFmpeg tasks, summed over L3 domains. This is occupancy, not hit ratio; use l3_cache_hit_ratio_pct for effectiveness.</span></div>",
        "<div class='legend-item'><strong>rdt_workload_total_mbm_gbps</strong><span>Memory bandwidth attributed to FFmpeg tasks by MBM. Unlike PCM DRAM columns, this is per-workload attribution rather than whole-worker traffic.</span></div>",
        "<div class='legend-item'><strong>rdt_noise_llc_occupancy_mib</strong><span>Cache footprint held by the noisy-neighbor tasks. Compare against the workload column to see how the two classes actually split the LLC, and to verify that a CAT profile took effect.</span></div>",
        "<div class='legend-item'><strong>rdt_noise_total_mbm_gbps</strong><span>Memory bandwidth attributed to the noisy-neighbor tasks. This is the direct check that an MBA profile throttled the noise class; a flat value means the throttle did not bite.</span></div>",
        "<div class='legend-item'><strong>rdt_encoder_* and rdt_decoder_*</strong><span>Separate RMID groups for encoder and decoder FFmpeg tasks, so each stage's LLC footprint and memory bandwidth can be read on its own. Workload columns are the sum of both roles.</span></div>",
        "<div class='legend-item'><strong>rdt_focus_session</strong><span>When set, RDT groups covered only that session, giving single-instance encoder and decoder attribution instead of all streams combined.</span></div>",
        "<div class='legend-item'><strong>l3_cache_hit_ratio_pct</strong><span>Whole-worker PCM L3 hits divided by hits plus misses over exact measurement window. Includes FFmpeg, noisy neighbor, Kubernetes, and all host work; not process-attributed. Use RDT columns for per-workload cache attribution.</span></div>",
        "<div class='legend-item'><strong>l3_cache_misses_per_second_millions</strong><span>Whole-worker PCM L3 misses per second in millions. Each miss becomes traffic toward DRAM or another cache, so rising misses with falling hit ratio indicate real LLC pressure.</span></div>",
        "<div class='legend-item'><strong>dram_read_gbps / dram_write_gbps</strong><span>Whole-worker PCM memory-controller read and write bandwidth over exact measurement window. Direct DRAM traffic evidence, host-wide, not per Pod.</span></div>",
        "<div class='legend-item'><strong>dram_total_gbps / dram_total_gbps_per_stream</strong><span>Sum of DRAM read and write bandwidth, and same value divided by benchmark stream count. Per-stream value is comparable only when the same worker runs no unrelated traffic.</span></div>",
        "<div class='legend-item'><strong>minimum_encoder_fps</strong><span>Lowest measurement-window average FPS among all encoder sessions. Pass requires at least configured threshold.</span></div>",
        "<div class='legend-item'><strong>sliced_threads</strong><span>1 = x264 slice-based threading; 0 = frame threading while slices still control output bitstream partitioning.</span></div>",
        "<div class='legend-item'><strong>result_directory</strong><span>Evidence directory containing exact configuration, metrics, commands, HTML, and XLSX report.</span></div>",
    ])
    profile_rows_html = "".join(
        "<tr>"
        + f"<td><strong>{html.escape(profile['profile'])}</strong></td>"
        + f"<td>{html.escape(profile['scope'])}</td>"
        + f"<td>{html.escape(profile['purpose'])}</td>"
        + f"<td><code>{html.escape(profile['args'])}</code></td>"
        + f"<td>{html.escape(profile['cpu_request'])}</td>"
        + f"<td>{html.escape(profile['memory_request'])}</td>"
        + f"<td>{html.escape(profile['memory_limit'])}</td>"
        + "</tr>"
        for profile in noise_profile_rows()
    )
    platform_headers = "".join(f"<th>{html.escape(label)}</th>" for _, label in platform_fields)
    rdt_profile_rows_html = "".join(
        "<tr>"
        + f"<td><strong>{html.escape(profile['profile'])}</strong></td>"
        + f"<td>{html.escape(profile['mechanism'])}</td>"
        + f"<td>{html.escape(profile['schemata'])}</td>"
        + f"<td>{html.escape(profile['effect'])}</td>"
        + f"<td>{html.escape(profile['use'])}</td>"
        + f"<td>{html.escape(profile['verify'])}</td>"
        + f"<td>{html.escape(profile['caveat'])}</td>"
        + "</tr>"
        for profile in rdt_profile_rows()
    )
    platform_rows_html = "".join(
        "<tr>" + "".join(
            f"<td>{html.escape(str(spec.get(field, 'unavailable')))}</td>"
            for field, _ in platform_fields
        ) + "</tr>"
        for spec in platform_by_node.values()
    ) or f"<tr><td colspan='{len(platform_fields)}'>No platform specification captured. Re-run tests to probe worker hardware.</td></tr>"
    stress_parameter_legend = "".join([
        "<div class='legend-item'><strong>--stream N</strong><span>Number of STREAM workers. Each one walks large arrays, so N sets how much memory traffic the neighbor generates. This is the knob that separates pod-a (4) from pod-b and pod-c (24).</span></div>",
        "<div class='legend-item'><strong>--stream-l3-size SIZE</strong><span>Working-set size per worker. Below the LLC size the pressure stays in cache; above it every access reaches DRAM. This is what separates pod-a (32M, cache pressure) from pod-c (512M, bandwidth pressure).</span></div>",
        "<div class='legend-item'><strong>--stream-index 0</strong><span>Sequential access. Predictable, prefetch-friendly traffic, so the measured bandwidth is reproducible rather than dominated by TLB misses.</span></div>",
        "<div class='legend-item'><strong>--timeout 0</strong><span>Run until stopped. The runner starts the neighbor before the warm-up and stops it after the measurement window, so the pressure covers the whole window.</span></div>",
        "<div class='legend-item'><strong>--metrics-brief / --log-file</strong><span>Report per-stressor totals on exit and keep the log inside the Pod, as evidence that the neighbor really ran.</span></div>",
        "<div class='legend-item'><strong>numactl --membind=&lt;socket&gt;</strong><span>Host scope only: binds the neighbor's memory to one socket's NUMA node, so each of the two host groups presses on its own memory controllers.</span></div>",
        "<div class='legend-item'><strong>Kubernetes requests / limits</strong><span>Guaranteed profiles set an integer CPU request equal to the limit, which is what makes the kubelet hand out exclusive cores. A neighbor with exclusive cores cannot steal FFmpeg's CPU time, so any damage it does is through cache and bandwidth alone.</span></div>",
    ])
    headers_html = "".join(
        "<th>" + html.escape(display_names[field]).replace("\n", "<br>") + "</th>"
        for field in fields
    )
    group_headers = "".join([
        "<th colspan='23'>Test configuration and RDT</th>",
        "<th colspan='1'>Throughput</th>",
        "<th colspan='3'>Encoder and host CPU</th>",
        "<th colspan='3'>Decoder and host CPU</th>",
        "<th colspan='2'>Cross-NUMA memory</th>",
        "<th colspan='6'>Cache and DRAM traffic</th>",
        "<th colspan='2'>DRAM headroom</th>",
        "<th colspan='1'>Result</th>",
        "<th colspan='5'>Execution details</th>",
        "<th colspan='1'>Evidence</th>",
    ])
    html_path.write_text(f"""<!doctype html><html><head><meta charset='utf-8'><title>MXL test summary</title>
<style>
:root{{--navy:#123b58;--navy2:#1e5d7b;--cyan:#20a4c7;--line:#d9e5ec;--soft:#f4f8fb;--green:#16834a;--red:#c43b4d;--ink:#172733;--muted:#687d8b}}
*{{box-sizing:border-box}}html{{background:#edf3f7}}body{{font:14px Inter,ui-sans-serif,system-ui,-apple-system,sans-serif;margin:0;color:var(--ink);background:radial-gradient(circle at 80% 0,#dceef5 0,transparent 28rem),#edf3f7}}
main{{max-width:1800px;margin:auto;min-height:100vh;padding:2rem 3rem 3rem}}h1,h2{{color:var(--navy)}}h1{{font-size:2rem;letter-spacing:-.035em;margin:0}}h2{{font-size:1.25rem;margin:2rem 0 1rem}}.hero{{position:relative;overflow:hidden;color:white;background:linear-gradient(120deg,#102f47,#17627f);border-radius:16px;padding:1.75rem 2rem;box-shadow:0 14px 36px #123b5828}}.hero:after{{content:'';position:absolute;right:-60px;top:-90px;width:300px;height:300px;border:55px solid #ffffff12;border-radius:50%}}.hero h1{{color:white}}.subtitle{{color:#c7dce7;margin:.5rem 0 0}}.eyebrow{{color:#76d4e9;font-size:.72rem;font-weight:800;letter-spacing:.16em;text-transform:uppercase;margin-bottom:.45rem}}
.cards{{display:grid;grid-template-columns:repeat(auto-fit,minmax(190px,1fr));gap:1rem;margin:1.25rem 0 2rem}}.card{{position:relative;background:#fff;border:1px solid var(--line);border-radius:12px;padding:1rem 1.15rem;box-shadow:0 5px 18px #15384d10;overflow:hidden}}.card:before{{content:'';position:absolute;inset:0 auto 0 0;width:4px;background:var(--cyan)}}.card span{{display:block;color:var(--muted);text-transform:uppercase;font-weight:700;letter-spacing:.06em;font-size:.7rem}}.card strong{{display:block;color:var(--navy);font-size:1.5rem;line-height:1.25;margin-top:.35rem}}.card small{{display:block;color:var(--muted);margin-top:.3rem}}
.section-head{{display:flex;align-items:end;justify-content:space-between;gap:1rem;margin-bottom:.8rem}}.section-head h2{{margin:0}}.key{{display:flex;gap:.9rem;color:var(--muted);font-size:.8rem}}.key span{{display:flex;align-items:center;gap:.35rem}}.key i{{width:9px;height:9px;border-radius:50%;background:var(--green)}}.key .fail-dot{{background:var(--red)}}
.table-shell{{background:white;border:1px solid var(--line);border-radius:12px;box-shadow:0 8px 24px #15384d12;overflow:hidden}}.scroll{{overflow:auto;max-height:66vh;scrollbar-color:#9db0bb #edf3f6;scrollbar-width:thin}}table{{border-collapse:separate;border-spacing:0;width:100%;white-space:nowrap}}th,td{{padding:.68rem .75rem;border-bottom:1px solid #e3ebf0;text-align:left}}thead th{{position:sticky;color:white;z-index:3}}thead .groups th{{top:0;background:#0e3249;color:#8dd5e8;padding:.42rem .75rem;font-size:.66rem;text-transform:uppercase;letter-spacing:.1em;border-right:1px solid #ffffff16}}thead .labels th{{top:27px;background:#174f6c;font-size:.76rem;line-height:1.15;box-shadow:inset 0 -1px #ffffff1c}}thead .labels th:nth-child(7),tbody td:nth-child(7){{width:112px;min-width:112px;max-width:112px}}tbody tr{{transition:background .12s}}tbody tr.pass{{background:#f3fbf6}}tbody tr.fail{{background:#fff6f7}}tbody tr:hover{{background:#eaf4f8}}td{{font-size:.82rem}}td.numeric{{font-variant-numeric:tabular-nums;text-align:right}}.status{{display:inline-flex;align-items:center;gap:.35rem;padding:.22rem .55rem;border-radius:99px;font-size:.68rem;font-weight:800;letter-spacing:.04em}}.status i{{width:6px;height:6px;border-radius:50%;background:currentColor}}.status.ok{{color:#08753c;background:#d9f3e4}}.status.bad{{color:#b4233c;background:#fbe1e5}}.report-link{{display:inline-flex;align-items:center;gap:.35rem;color:#126c90;font-weight:700;text-decoration:none}}.report-link:hover{{color:#0a405d;text-decoration:underline}}
.legend{{display:grid;grid-template-columns:repeat(auto-fit,minmax(280px,1fr));gap:.8rem;margin-bottom:1.5rem}}.legend-item{{background:white;border:1px solid var(--line);border-radius:9px;padding:.8rem 1rem;box-shadow:0 3px 12px #15384d0b}}.legend-item strong,.legend-item span{{display:block}}.legend-item strong{{color:var(--navy);font-size:.82rem}}.legend-item span{{color:var(--muted);font-size:.8rem;line-height:1.45;margin-top:.25rem}}
.comparison{{overflow:auto;background:white;border:1px solid var(--line);border-radius:12px;box-shadow:0 5px 18px #15384d10}}.comparison table{{white-space:normal;min-width:1200px}}.comparison th{{position:static;background:#174f6c;color:white;font-size:.74rem}}.comparison td{{vertical-align:top;line-height:1.4;min-width:190px}}.comparison td:first-child{{min-width:170px;background:#f4f8fb}}.comparison td strong,.comparison td small{{display:block}}.comparison td strong{{color:var(--navy);font-size:.9rem}}.comparison td small{{color:var(--muted);margin-top:.3rem;line-height:1.35}}
.winners{{display:grid;grid-template-columns:repeat(auto-fit,minmax(235px,1fr));gap:1rem;margin-bottom:2rem}}.winner{{position:relative;background:white;border:1px solid var(--line);border-top:4px solid #6d91a4;border-radius:12px;padding:1rem 1.1rem;box-shadow:0 6px 18px #15384d10}}.winner.recommended{{border-color:#2a9d68;border-top-color:#16834a;background:linear-gradient(145deg,#f2fbf6,#fff)}}.winner-badge{{position:absolute;right:.8rem;top:.7rem;color:#08753c;background:#d9f3e4;border-radius:99px;padding:.22rem .5rem;font-size:.65rem;font-weight:800;text-transform:uppercase;letter-spacing:.05em}}.winner-name{{color:var(--navy);font-size:1rem;font-weight:800;text-transform:uppercase;letter-spacing:.04em}}.winner-score{{display:flex;align-items:baseline;gap:.5rem;margin:.8rem 0}}.winner-score strong{{font-size:2.25rem;line-height:1;color:var(--navy)}}.winner-score span{{color:var(--muted);font-size:.78rem}}.winner dl{{margin:0;border-top:1px solid var(--line)}}.winner dl div{{display:flex;justify-content:space-between;gap:1rem;padding:.38rem 0;border-bottom:1px solid #e8eef2}}.winner dt{{color:var(--muted);font-size:.75rem}}.winner dd{{margin:0;font-weight:700;font-size:.78rem;text-align:right}}.winner-link{{margin-top:.8rem}}
@media(max-width:800px){{main{{padding:1rem}}.hero{{padding:1.4rem}}h1{{font-size:1.55rem}}.section-head{{align-items:start;flex-direction:column}}}}
</style></head><body><main>
<header class='hero'><div class='eyebrow'>Performance laboratory</div><h1>FFmpeg + MXL capacity summary</h1><p class='subtitle'>Passing runs ranked by stream count, then minimum encoder FPS.</p></header>
<div class='cards'><div class='card'><span>Best passing streams</span><strong>{html.escape(best_streams)}</strong><small>highest strict-pass density</small></div><div class='card'><span>Best scenario</span><strong>{html.escape(best_scenario)}</strong><small>top-ranked passing profile</small></div><div class='card'><span>Required encoder FPS</span><strong>{min_fps:.3f}</strong><small>minimum for every stream</small></div></div>
<h2>Tested worker platform specification</h2><p class='section-note'>Probed from each worker during runs. Theoretical DRAM peak is populated channels x memory transfer rate x 8 bytes. It is a hardware ceiling derived from configuration, not a measured or achievable rate; real sustained bandwidth stays well below it. Unavailable transfer rate means DMI was unreadable and no <code>&lt;NODE&gt;_MEM_TRANSFER_MT_S</code> value was declared in <code>config/nodes.env</code>.</p>
<div class='comparison'><table><thead><tr>{platform_headers}</tr></thead><tbody>{platform_rows_html}</tbody></table></div>
<div class='section-head'><div><h2>Best result by scenario</h2><p class='section-note'>Ranking: highest strict-pass stream count first; lowest measured total FFmpeg CPU demand breaks stream-count ties; minimum encoder FPS breaks remaining ties. Best overall option highlighted.</p></div></div>
<div class='winners'>{winner_cards}</div>
<div class='section-head'><h2>Benchmark results — no noise</h2><div class='key'><span><i></i>Pass</span><span><i class='fail-dot'></i>Fail</span></div></div>
<div class='table-shell'><div class='scroll'><table><thead><tr class='groups'>{group_headers}</tr><tr class='labels'>{headers_html}</tr></thead><tbody>{standard_trs}</tbody></table></div></div>
<div class='section-head'><h2>Noisy-neighbor benchmark results</h2><div class='key'><span><i></i>Pass</span><span><i class='fail-dot'></i>Fail</span></div></div>
<div class='table-shell'><div class='scroll'><table><thead><tr class='groups'>{group_headers}</tr><tr class='labels'>{headers_html}</tr></thead><tbody>{noise_trs}</tbody></table></div></div>
<h2>Scenario comparison</h2><div class='comparison'><table><thead><tr><th>Scenario</th><th>CPU placement</th><th>Codec and threading</th><th>CPU allocation</th><th>Purpose</th><th>Main trade-off</th></tr></thead><tbody>{scenario_rows}</tbody></table></div>
<h2>Metric and column legend</h2><div class='legend'>{metric_legend}</div>
<h2>Noisy-neighbor scenario reference</h2><p>Exact profile parameters used by this checkout. The documented pressure levels are targets; the achieved CPU, LLC and DRAM figures must be read from each run's own evidence. <strong>host-a is not a Pod:</strong> it runs outside kubepods, so it can use CPUs the kubelet gave exclusively to pinned FFmpeg containers and its results include direct CPU contention on top of cache and bandwidth contention. <strong>pod-a, pod-b and pod-c are Guaranteed Pods</strong> with exclusive cores, one per socket, so they contend only through the shared LLC, memory controllers and interconnect.</p>
<div class='comparison'><table><thead><tr><th>Profile</th><th>Scope</th><th>Purpose</th><th>stress-ng arguments</th><th>CPU request</th><th>Memory request</th><th>Memory limit</th></tr></thead><tbody>{profile_rows_html}</tbody></table></div>
<h2>stress-ng parameter reference</h2><div class='legend'>{stress_parameter_legend}</div>
<h2>RDT control profile reference</h2><p>Exact meaning of every value in the <strong>RDT control profile</strong> column. RDT manages only two shared resources: LLC capacity through CAT and memory bandwidth through MBA. It never manages CPU time, frequency, or UPI links, so host-scope noise that shares workload CPUs stays partly unmitigated. Control is applied to the noise class; nothing is reserved for FFmpeg.</p>
<p><strong>How a profile is applied.</strong> Before the measurement window the run resolves Pod UIDs to host processes, expands them to thread IDs, and writes those IDs into resctrl groups: <code>encoder</code> and <code>decoder</code> for FFmpeg, <code>noise</code> for stress-ng. With <code>none</code> the groups live under <code>mon_groups/</code> and only take an RMID. With any control profile they are full control groups that also take a CLOS, and the schemata shown below is written to each one. Every group except <code>noise</code> receives the protected allocation. Association is per task, so a control profile still works where noise and FFmpeg share the same CPUs. Counters are sampled once per second for the whole window, then the groups are removed and every task returns to the default class; <code>rdt-after.json</code> records the restored state.</p>
<p><strong>Reading the numbers.</strong> CMT occupancy is a footprint in bytes, not a hit ratio, and MBM is bandwidth attributed to a group, not the whole worker. Use the PCM columns for host totals and the RDT columns for attribution. Absolute MBM can differ from PCM by a wide margin on some traffic patterns, so trust the relative change between rows of the same A/B set. A group that reports zero occupancy and zero bandwidth for the entire window is treated as a failure, because it means the tasks were never associated.</p>
<div class='comparison'><table><thead><tr><th>Profile</th><th>RDT mechanism</th><th>Applied schemata</th><th>What it does</th><th>When to use</th><th>How to verify it worked</th><th>Limits and honest caveat</th></tr></thead><tbody>{rdt_profile_rows_html}</tbody></table></div>
</main></body></html>""")
    return html_path, xlsx_path
