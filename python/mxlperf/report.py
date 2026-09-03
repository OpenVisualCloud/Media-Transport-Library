from __future__ import annotations

import csv
import html
import json
from pathlib import Path

import yaml

from .platform import platform_summary_line
from .render import resolved_ffmpeg_command
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def read_csv(path: Path) -> list[dict[str, str]]:
    if not path.is_file():
        return []
    with path.open(newline="") as handle:
        return list(csv.DictReader(handle))


def autosize(sheet) -> None:
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column in range(1, sheet.max_column + 1):
        width = max(len(str(sheet.cell(row, column).value or "")) for row in range(1, sheet.max_row + 1))
        sheet.column_dimensions[get_column_letter(column)].width = min(80, max(12, width + 2))
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")


def aggregate_value(value: str, field: str = "avg") -> float | None:
    try:
        parts = dict(item.split("=", 1) for item in value.split(";"))
        return float(parts[field])
    except (ValueError, KeyError):
        return None


def aggregate_metrics(
    rows: list[dict[str, str]], metric: str, *, role: str = "", category: str = ""
) -> list[float]:
    values = []
    for row in rows:
        if row.get("metric") != metric:
            continue
        if role and row.get("role") != role:
            continue
        if category and row.get("category") != category:
            continue
        value = aggregate_value(row.get("value", ""))
        if value is not None:
            values.append(value)
    return values


def plain_metrics(rows: list[dict[str, str]], metric: str, *, role: str = "") -> list[float]:
    values = []
    for row in rows:
        if row.get("metric") != metric or (role and row.get("role") != role):
            continue
        try:
            values.append(float(row["value"]))
        except ValueError:
            pass
    return values


def numa_bandwidth_rows(rows: list[dict[str, str]]) -> list[dict[str, object]]:
    sockets: dict[int, dict[str, list[float]]] = {}
    metric_roles = {
        "socket_local_memory_bandwidth_bytes_per_second": "local",
        "socket_remote_memory_bandwidth_bytes_per_second": "remote",
        "socket_upi_incoming_bytes_per_second": "upi",
    }
    for row in rows:
        kind = metric_roles.get(row.get("metric", ""))
        if not kind:
            continue
        try:
            labels = json.loads(row.get("scope", ""))
            socket = int(labels["socket"])
        except (json.JSONDecodeError, KeyError, TypeError, ValueError):
            continue
        value = aggregate_value(row.get("value", ""))
        if value is not None:
            sockets.setdefault(socket, {"local": [], "remote": [], "upi": []})[kind].append(value)
    result = []
    for socket, values in sorted(sockets.items()):
        local = sum(values["local"]) / len(values["local"]) if values["local"] else 0.0
        remote = sum(values["remote"]) / len(values["remote"]) if values["remote"] else 0.0
        upi = sum(values["upi"]) / len(values["upi"]) if values["upi"] else 0.0
        total = local + remote
        result.append({
            "Socket": socket,
            "Local memory bandwidth (GB/s)": round(local / 1e9, 3),
            "Remote memory bandwidth (GB/s)": round(remote / 1e9, 3),
            "Remote memory ratio (%)": round(remote * 100 / total, 3) if total else 0.0,
            "UPI incoming bandwidth (GB/s)": round(upi / 1e9, 3),
        })
    return result


def add_sheet(workbook: Workbook, name: str, rows: list[dict[str, str]]) -> None:
    sheet = workbook.create_sheet(name[:31])
    if not rows:
        sheet.append(["No data"]); return
    fields = list(rows[0])
    sheet.append(fields)
    for row in rows:
        sheet.append([row.get(field, "") for field in fields])
    autosize(sheet)


def per_core_usage(
    rows: list[dict[str, str]], duration: float, system_rows: list[dict[str, str]] | None = None
) -> list[dict[str, object]]:
    real_usage = {}
    for row in system_rows or []:
        try:
            # real_core_total_usage_pct is the pre-rename column name, still present
            # in result directories collected before it.
            usage = row.get("real_cpu_total_usage_pct") or row["real_core_total_usage_pct"]
            real_usage[int(row["cpu_id"])] = float(usage)
        except (KeyError, ValueError):
            continue
    cores: dict[int, dict[str, float]] = {}
    for row in rows:
        try:
            cpu_id = int(row["cpu_id"])
            seconds = float(row["ffmpeg_cpu_seconds"])
        except (KeyError, ValueError):
            continue
        values = cores.setdefault(cpu_id, {"encoder": 0.0, "decoder": 0.0})
        role = row.get("role", "")
        if role in values:
            values[role] += seconds
    elapsed = max(0.001, duration)
    result = []
    for cpu_id, values in sorted(cores.items()):
        encoder_seconds = values["encoder"]
        decoder_seconds = values["decoder"]
        total_seconds = encoder_seconds + decoder_seconds
        result.append({
            "CPU ID": cpu_id,
            "FFmpeg avg usage (%)": round(total_seconds * 100 / elapsed, 3),
            "Encoder avg usage (%)": round(encoder_seconds * 100 / elapsed, 3),
            "Real CPU total usage (%)": round(real_usage[cpu_id], 3) if cpu_id in real_usage else "unavailable",
            "Decoder avg usage (%)": round(decoder_seconds * 100 / elapsed, 3),
            "FFmpeg CPU time (s)": round(total_seconds, 3),
            "Encoder CPU time (s)": round(encoder_seconds, 3),
            "Decoder CPU time (s)": round(decoder_seconds, 3),
        })
    return result


def build_report(result: Path) -> None:
    metrics = read_csv(result / "metrics.csv")
    placement = read_csv(result / "cpu-placement.csv")
    usage = read_csv(result / "cpu-usage.csv")
    system_usage = read_csv(result / "core-system-usage.csv")
    numa_counters = read_csv(result / "numa-node-counters.csv")
    rdt_summary = json.loads((result / "rdt-summary.json").read_text()) if (result / "rdt-summary.json").is_file() else []
    rdt_stop = json.loads((result / "rdt-stop.json").read_text()) if (result / "rdt-stop.json").is_file() else {}
    config = json.loads((result / "config.json").read_text()) if (result / "config.json").is_file() else {}
    host = json.loads((result / "host.json").read_text()) if (result / "host.json").is_file() else {}
    window = json.loads((result / "window.json").read_text()) if (result / "window.json").is_file() else {}
    noise_path = result / "noisy-neighbor.json"
    noise_before_path = result / "noisy-neighbor-before.json"
    noise_after_path = result / "noisy-neighbor-after.json"
    noise_log_path = result / "noisy-neighbor.log"
    # Keep old result directories reportable after spelling correction.
    noise_path = noise_path if noise_path.is_file() else result / "noisy-neighbour.json"
    noise_before_path = noise_before_path if noise_before_path.is_file() else result / "noisy-neighbour-before.json"
    noise_after_path = noise_after_path if noise_after_path.is_file() else result / "noisy-neighbour-after.json"
    noise_log_path = noise_log_path if noise_log_path.is_file() else result / "noisy-neighbour.log"
    noise = json.loads(noise_path.read_text()) if noise_path.is_file() else {"enabled": False}
    noise_before = json.loads(noise_before_path.read_text()) if noise_before_path.is_file() else {}
    noise_after = json.loads(noise_after_path.read_text()) if noise_after_path.is_file() else {}
    noise_log = noise_log_path.read_text() if noise_log_path.is_file() else ""
    noise_numa = noise_after.get("numa", {}) if isinstance(noise_after.get("numa"), dict) else {}
    noise_numa_precheck = noise_numa.get("precheck", {}) if isinstance(noise_numa.get("precheck"), dict) else {}
    noise_numa_status = str(noise_numa.get("status", "n/a"))
    same_socket = noise_numa.get("same_socket")
    same_socket_label = "yes" if same_socket is True else "no" if same_socket is False else "n/a"
    precheck_status = str(noise_numa_precheck.get("status", "n/a")) if noise_numa_precheck else "n/a"

    def cgroup_value(snapshot: dict, key: str) -> int | None:
        raw = snapshot.get("runtime_evidence", "")
        values: list[int] = []
        for line in raw.splitlines():
            parts = line.split()
            if len(parts) == 2 and parts[0] == key and parts[1].isdigit():
                values.append(int(parts[1]))
        return sum(values) if values else None

    noise_cpu_seconds = None
    before_usage = cgroup_value(noise_before, "usage_usec")
    after_usage = cgroup_value(noise_after, "usage_usec")
    if before_usage is not None and after_usage is not None:
        noise_cpu_seconds = max(0, after_usage - before_usage) / 1e6
    measurement_duration = float(window.get("duration_s", 0) or 0)
    core_usage = per_core_usage(usage, measurement_duration, system_usage)
    for row in metrics:
        if row.get("role"):
            continue
        try:
            labels = json.loads(row.get("scope", ""))
        except (json.JSONDecodeError, TypeError):
            continue
        pod = labels.get("pod", labels.get("exported_pod", ""))
        row["role"] = "decoder" if "-decoder-" in pod else "encoder" if "-encoder-" in pod else ""
    command_path = result / "ffmpeg-commandlines.json"
    if command_path.is_file():
        command_lines = json.loads(command_path.read_text())
    else:
        command_lines = []
        workload_path = result / "workload.yaml"
        if workload_path.is_file():
            for document in yaml.safe_load_all(workload_path.read_text()):
                if not document or document.get("kind") != "Pod":
                    continue
                if document["metadata"].get("labels", {}).get("role") in ("noisy-neighbor", "noisy-neighbour"):
                    continue
                container = document["spec"]["containers"][0]
                command_lines.append({
                    "pod": document["metadata"]["name"],
                    "role": document["metadata"]["labels"]["role"],
                    "command": resolved_ffmpeg_command(container),
                })

    threshold = float(window.get("min_fps", config.get("LAB_MIN_FPS", 59.5)))
    encoder_fps = aggregate_metrics(metrics, "fps", role="encoder", category="FFmpeg progress")
    decoder_fps = aggregate_metrics(metrics, "fps", role="decoder", category="FFmpeg progress")
    node_busy = aggregate_metrics(metrics, "node_busy_cores")
    encoder_cpu = plain_metrics(metrics, "FFmpeg CPU demand", role="encoder")
    decoder_cpu = plain_metrics(metrics, "FFmpeg CPU demand", role="decoder")
    encoder_util = plain_metrics(metrics, "average FFmpeg utilization per measured used CPU", role="encoder")
    encoder_memory = aggregate_metrics(metrics, "workload_memory_bytes", role="encoder")
    decoder_memory = aggregate_metrics(metrics, "workload_memory_bytes", role="decoder")
    l3_hit_ratio = aggregate_metrics(metrics, "l3_cache_hit_ratio")
    l3_hits = aggregate_metrics(metrics, "l3_cache_hits_per_second")
    l3_misses = aggregate_metrics(metrics, "l3_cache_misses_per_second")
    dram_read = aggregate_metrics(metrics, "dram_read_bytes_per_second")
    dram_write = aggregate_metrics(metrics, "dram_write_bytes_per_second")
    cross_numa_upi = aggregate_metrics(metrics, "cross_numa_upi_incoming_bytes_per_second")
    stream_count = int(config.get("STREAMS", 0))
    platform_spec = host.get("platform_spec", {}) if isinstance(host.get("platform_spec"), dict) else {}
    dram_total_gbps = (
        (sum(dram_read) / len(dram_read) + sum(dram_write) / len(dram_write)) / 1e9
        if dram_read and dram_write else None
    )
    theoretical_peak_gbps = platform_spec.get("theoretical_dram_gbps")
    dram_peak_pct = (
        dram_total_gbps / theoretical_peak_gbps * 100
        if dram_total_gbps is not None and theoretical_peak_gbps else None
    )
    cross_numa_upi_gbps = sum(cross_numa_upi) / len(cross_numa_upi) / 1e9 if cross_numa_upi else None
    cross_numa_upi_per_stream = cross_numa_upi_gbps / stream_count if cross_numa_upi_gbps is not None and stream_count else None
    socket_numa_bandwidth = numa_bandwidth_rows(metrics)
    numa_deltas: dict[str, int] = {}
    for row in numa_counters:
        try:
            numa_deltas[row["counter"]] = numa_deltas.get(row["counter"], 0) + int(row["delta_pages"])
        except (KeyError, ValueError):
            continue
    encoder_min = min(encoder_fps) if encoder_fps else None
    decoder_min = min(decoder_fps) if decoder_fps else None
    passed = encoder_min is not None and len(encoder_fps) == int(config.get("STREAMS", 0)) and encoder_min >= threshold
    status = "PASS" if passed else "FAIL"
    average_encoder_util = sum(encoder_util) / len(encoder_util) if encoder_util else None
    if passed:
        interpretation = "Target met; unused encoder CPU is expected because input is paced at 60 FPS"
    elif average_encoder_util is not None and average_encoder_util < 85:
        interpretation = "Target missed without CPU saturation; codec parallelism or thread/slice/core balance limits throughput"
    else:
        interpretation = "Target missed near CPU saturation; add encoder CPUs or use faster preset"
    headline = [
        {"group": "Result", "field": "Result", "value": status},
        {"group": "Platform", "field": "Worker platform", "value": platform_summary_line(platform_spec)},
        {"group": "Test", "field": "Scenario", "value": config.get("SCENARIO", "")},
        {"group": "Test", "field": "CPU placement", "value": config.get("PLACEMENT", "")},
        {"group": "Test", "field": "Worker node", "value": config.get("NODE", config.get("LAB_DEFAULT_NODE", ""))},
        {"group": "Test", "field": "Streams", "value": config.get("STREAMS", "")},
        {"group": "Noise", "field": "Noisy neighbor", "value": noise.get("profile", "disabled") if noise.get("enabled") else "disabled"},
        {"group": "Noise", "field": "stress-ng image", "value": noise.get("image", "n/a")},
        {"group": "Noise", "field": "stress-ng arguments", "value": " ".join(noise.get("args", [])) or "n/a"},
        {"group": "Noise", "field": "Measured stress-ng cgroup CPU time", "value": f"{noise_cpu_seconds:.3f} seconds" if noise_cpu_seconds is not None else "n/a"},
        {"group": "Noise", "field": "Final Pod phase", "value": noise_after.get("phase", "n/a")},
        {"group": "Noise", "field": "NN socket placement status", "value": noise_numa_status},
        {"group": "Noise", "field": "NN pods on same socket", "value": same_socket_label},
        {"group": "Noise", "field": "NN precheck socket status", "value": precheck_status},
        {"group": "RDT", "field": "Monitoring", "value": "enabled" if config.get("RDT_MONITOR") == "1" or config.get("RDT_CONTROL_PROFILE", "none") != "none" else "disabled"},
        {"group": "RDT", "field": "Control profile", "value": config.get("RDT_CONTROL_PROFILE", "none")},
        {"group": "RDT", "field": "State restored", "value": rdt_stop.get("restored", "n/a")},
        {"group": "Codec", "field": "Preset", "value": config.get("PRESET", "")},
        {"group": "Codec", "field": "Threads / slices", "value": f"{config.get('ENC_THREADS', 'auto')} / {config.get('SLICES', 'auto')}"},
        {"group": "Throughput", "field": "Minimum encoder FPS", "value": f"{encoder_min:.3f}" if encoder_min is not None else "missing"},
        {"group": "Throughput", "field": "Minimum decoder FPS", "value": f"{decoder_min:.3f}" if decoder_min is not None else "missing"},
        {"group": "Throughput", "field": "Required FPS", "value": f"{threshold:.3f}"},
        {"group": "CPU", "field": "Whole worker CPU busy", "value": f"{sum(node_busy) / len(node_busy):.3f} core-equivalents" if node_busy else "missing"},
        {"group": "CPU", "field": "FFmpeg encoder CPU demand", "value": f"{sum(encoder_cpu):.3f} core-equivalents" if encoder_cpu else "missing"},
        {"group": "CPU", "field": "Encoder average utilization per distinct used CPU", "value": f"{average_encoder_util:.1f}%" if average_encoder_util is not None else "missing"},
        {"group": "CPU", "field": "FFmpeg decoder CPU demand", "value": f"{sum(decoder_cpu):.3f} core-equivalents" if decoder_cpu else "missing"},
        {"group": "Memory", "field": "Encoder containers working-set memory", "value": f"{sum(encoder_memory)/2**20:.1f} MiB" if encoder_memory else "missing"},
        {"group": "Memory", "field": "Decoder containers working-set memory", "value": f"{sum(decoder_memory)/2**20:.1f} MiB" if decoder_memory else "missing"},
        {"group": "Hardware", "field": "Whole worker L3 cache hit ratio", "value": f"{sum(l3_hit_ratio)/len(l3_hit_ratio)*100:.1f}%" if l3_hit_ratio else "missing"},
        {"group": "Hardware", "field": "Whole worker DRAM read bandwidth", "value": f"{sum(dram_read)/len(dram_read)/1e9:.2f} GB/s" if dram_read else "missing"},
        {"group": "Hardware", "field": "Whole worker DRAM write bandwidth", "value": f"{sum(dram_write)/len(dram_write)/1e9:.2f} GB/s" if dram_write else "missing"},
        {"group": "Hardware", "field": "Whole worker DRAM read+write bandwidth", "value": f"{dram_total_gbps:.2f} GB/s" if dram_total_gbps is not None else "missing"},
        {"group": "Hardware", "field": "Theoretical DRAM peak bandwidth", "value": f"{theoretical_peak_gbps:.1f} GB/s" if theoretical_peak_gbps else "unavailable"},
        {"group": "Hardware", "field": "Measured DRAM share of theoretical peak", "value": f"{dram_peak_pct:.1f}%" if dram_peak_pct is not None else "unavailable"},
        {"group": "Hardware", "field": "Whole worker DRAM read+write bandwidth", "value": f"{dram_total_gbps:.2f} GB/s" if dram_total_gbps is not None else "missing"},
        {"group": "Hardware", "field": "Theoretical DRAM peak bandwidth", "value": f"{theoretical_peak_gbps:.1f} GB/s" if theoretical_peak_gbps else "unavailable"},
        {"group": "Hardware", "field": "Measured DRAM share of theoretical peak", "value": f"{dram_peak_pct:.1f}%" if dram_peak_pct is not None else "unavailable"},
        {"group": "NUMA", "field": "Cross-socket UPI incoming bandwidth", "value": f"{cross_numa_upi_gbps:.2f} GB/s" if cross_numa_upi_gbps is not None else "missing"},
        {"group": "NUMA", "field": "Cross-socket UPI incoming per stream", "value": f"{cross_numa_upi_per_stream:.3f} GB/s" if cross_numa_upi_per_stream is not None else "missing"},
        {"group": "NUMA", "field": "NUMA miss delta", "value": f"{numa_deltas['numa_miss']} pages" if "numa_miss" in numa_deltas else "missing"},
        {"group": "NUMA", "field": "Other-node allocation delta", "value": f"{numa_deltas['other_node']} pages" if "other_node" in numa_deltas else "missing"},
        {"group": "Analysis", "field": "Interpretation", "value": interpretation},
        {"group": "Test", "field": "Measurement window", "value": f"{window.get('duration_s', '')} seconds"},
    ]

    platform_rows = [
        {"Field": "CPU model", "Value": platform_spec.get("cpu_model", "unavailable")},
        {"Field": "Sockets", "Value": platform_spec.get("sockets", "unavailable")},
        {"Field": "Cores per socket", "Value": platform_spec.get("cores_per_socket", "unavailable")},
        {"Field": "Threads per core", "Value": platform_spec.get("threads_per_core", "unavailable")},
        {"Field": "Logical CPUs", "Value": platform_spec.get("logical_cpus", "unavailable")},
        {"Field": "NUMA nodes", "Value": platform_spec.get("numa_nodes", "unavailable")},
        {"Field": "CPU max frequency (MHz)", "Value": platform_spec.get("cpu_max_mhz", "unavailable")},
        {"Field": "L3 cache", "Value": platform_spec.get("l3_cache", "unavailable")},
        {"Field": "Installed memory (GiB)", "Value": platform_spec.get("memory_total_gib", "unavailable")},
        {"Field": "Populated DIMMs", "Value": platform_spec.get("populated_dimms", "unavailable")},
        {"Field": "DIMM size (GiB) and type", "Value": f"{platform_spec.get('dimm_size_gib', 'unavailable')} {platform_spec.get('dimm_type', '')}".strip()},
        {"Field": "Populated memory channels", "Value": platform_spec.get("populated_channels", "unavailable")},
        {"Field": "Memory channels per socket", "Value": platform_spec.get("channels_per_socket", "unavailable")},
        {"Field": "Memory transfer rate (MT/s)", "Value": platform_spec.get("memory_transfer_mt_s", "unavailable")},
        {"Field": "Memory transfer rate source", "Value": platform_spec.get("memory_transfer_source", "") or "unavailable"},
        {"Field": "Theoretical DRAM peak bandwidth (GB/s)", "Value": theoretical_peak_gbps or "unavailable"},
        {"Field": "Measured DRAM read+write (GB/s)", "Value": round(dram_total_gbps, 2) if dram_total_gbps is not None else "unavailable"},
        {"Field": "Measured share of theoretical peak (%)", "Value": round(dram_peak_pct, 1) if dram_peak_pct is not None else "unavailable"},
    ]

    workbook = Workbook()
    workbook.remove(workbook.active)
    add_sheet(workbook, "Summary", headline)
    summary_sheet = workbook["Summary"]
    summary_sheet.sheet_view.showGridLines = False
    summary_sheet["C2"].font = Font(bold=True, color="FFFFFF", size=14)
    summary_sheet["C2"].fill = PatternFill("solid", fgColor="22863A" if passed else "CB2431")
    summary_sheet.row_dimensions[2].height = 26
    group_colors = {"Result": "D9EAD3", "Test": "DCEAF4", "Noise": "F4CCCC", "RDT": "D9EAD3", "Codec": "EDE2F6", "Throughput": "D9EAD3", "CPU": "FCE5CD", "Memory": "D9EAD3", "Hardware": "FFF2CC", "NUMA": "DDEBF7", "Analysis": "EDE2F6"}
    for row in range(2, summary_sheet.max_row + 1):
        group = str(summary_sheet.cell(row, 1).value)
        summary_sheet.cell(row, 1).font = Font(bold=True, color="1F4E78")
        summary_sheet.cell(row, 2).font = Font(bold=True, color="1F4E78")
        fill = PatternFill("solid", fgColor=group_colors.get(group, "DCEAF4"))
        summary_sheet.cell(row, 1).fill = fill
        summary_sheet.cell(row, 2).fill = fill
        for column in (1, 2, 3):
            summary_sheet.cell(row, column).alignment = Alignment(vertical="center")
            summary_sheet.cell(row, column).border = Border(bottom=Side(style="thin", color="D9E2E9"))
    add_sheet(workbook, "Metrics", metrics)
    add_sheet(workbook, "CPU placement", placement)
    add_sheet(workbook, "Platform specification", platform_rows)
    add_sheet(workbook, "Core CPU usage", core_usage)
    add_sheet(workbook, "NUMA bandwidth", socket_numa_bandwidth)
    add_sheet(workbook, "NUMA counters", numa_counters)
    add_sheet(workbook, "RDT attribution", rdt_summary)
    add_sheet(workbook, "Per process CPU detail", [{k: v for k, v in row.items() if k != "attribution"} for row in usage])
    add_sheet(workbook, "FFmpeg commands", command_lines)
    add_sheet(workbook, "Noisy neighbor", [
        {"field": "enabled", "value": noise.get("enabled", False)},
        {"field": "profile", "value": noise.get("profile", "disabled")},
        {"field": "image", "value": noise.get("image", "")},
        {"field": "arguments", "value": " ".join(noise.get("args", []))},
        {"field": "CPU time during measurement (s)", "value": noise_cpu_seconds if noise_cpu_seconds is not None else "n/a"},
        {"field": "final phase", "value": noise_after.get("phase", "n/a")},
        {"field": "socket placement status", "value": noise_numa_status},
        {"field": "pods on same socket", "value": same_socket_label},
        {"field": "precheck socket status", "value": precheck_status},
        {"field": "socket details", "value": json.dumps(noise_numa.get("pods", [])) if noise_numa else "n/a"},
        {"field": "runtime evidence after", "value": noise_after.get("runtime_evidence", "")},
        {"field": "log", "value": noise_log},
    ])
    add_sheet(workbook, "Configuration", [{"key": k, "value": v} for k, v in sorted(config.items())])
    add_sheet(workbook, "Host", [{"key": k, "value": json.dumps(v) if isinstance(v, (dict, list)) else v} for k, v in host.items()])
    workbook.save(result / "report.xlsx")

    def table(rows: list[dict[str, str]]) -> str:
        if not rows:
            return "<p>No data.</p>"
        fields = list(rows[0])
        head = "".join(f"<th>{html.escape(field)}</th>" for field in fields)
        body = "".join("<tr>" + "".join(f"<td>{html.escape(str(row.get(field, '')))}</td>" for field in fields) + "</tr>" for row in rows)
        return f"<div class='scroll'><table><thead><tr>{head}</tr></thead><tbody>{body}</tbody></table></div>"

    throughput_cards = [
        ("Result", status, "pass" if passed else "fail"),
        ("Streams", str(config.get("STREAMS", "")), ""),
        ("Min encoder FPS", f"{encoder_min:.3f}" if encoder_min is not None else "missing", ""),
        ("Required FPS", f"{threshold:.3f}", ""),
        ("Min decoder FPS", f"{decoder_min:.3f}" if decoder_min is not None else "missing", ""),
    ]
    cpu_cards = [
        ("Whole worker CPU busy", f"{sum(node_busy) / len(node_busy):.3f} core-equivalents" if node_busy else "missing", ""),
        ("FFmpeg encoder CPU demand", f"{sum(encoder_cpu):.3f} core-equivalents" if encoder_cpu else "missing", ""),
        ("Encoder avg utilization / distinct used CPU", f"{average_encoder_util:.1f}%" if average_encoder_util is not None else "missing", ""),
        ("FFmpeg decoder CPU demand", f"{sum(decoder_cpu):.3f} core-equivalents" if decoder_cpu else "missing", ""),
    ]
    memory_cards = [
        ("Encoder containers working set", f"{sum(encoder_memory)/2**20:.1f} MiB" if encoder_memory else "missing", ""),
        ("Decoder containers working set", f"{sum(decoder_memory)/2**20:.1f} MiB" if decoder_memory else "missing", ""),
        ("Whole worker DRAM read", f"{sum(dram_read)/len(dram_read)/1e9:.2f} GB/s" if dram_read else "missing", ""),
        ("Whole worker DRAM write", f"{sum(dram_write)/len(dram_write)/1e9:.2f} GB/s" if dram_write else "missing", ""),
        ("Theoretical DRAM peak", f"{theoretical_peak_gbps:.1f} GB/s" if theoretical_peak_gbps else "unavailable", ""),
        ("Measured share of DRAM peak", f"{dram_peak_pct:.1f}%" if dram_peak_pct is not None else "unavailable", ""),
    ]
    cache_cards = [
        ("Whole worker L3 hit ratio", f"{sum(l3_hit_ratio)/len(l3_hit_ratio)*100:.1f}%" if l3_hit_ratio else "missing", ""),
        ("Whole worker L3 hits", f"{sum(l3_hits)/len(l3_hits)/1e6:.2f} million/s" if l3_hits else "missing", ""),
        ("Whole worker L3 misses", f"{sum(l3_misses)/len(l3_misses)/1e6:.2f} million/s" if l3_misses else "missing", ""),
    ]
    numa_cards = [
        ("Cross-socket UPI incoming bandwidth", f"{cross_numa_upi_gbps:.2f} GB/s" if cross_numa_upi_gbps is not None else "missing", ""),
        ("Cross-socket UPI incoming per stream", f"{cross_numa_upi_per_stream:.3f} GB/s" if cross_numa_upi_per_stream is not None else "missing", ""),
        ("NUMA miss delta", f"{numa_deltas['numa_miss']} pages" if "numa_miss" in numa_deltas else "missing", ""),
        ("Other-node allocation delta", f"{numa_deltas['other_node']} pages" if "other_node" in numa_deltas else "missing", ""),
    ]
    def render_cards(cards):
        return "".join(
        f"<div class='card {css}'><span>{html.escape(label)}</span><strong>{html.escape(value)}</strong></div>"
        for label, value, css in cards
        )
    def core_usage_table(rows: list[dict[str, object]]) -> str:
        if not rows:
            return "<p>No per-core FFmpeg CPU data.</p>"
        def real_usage_text(row: dict[str, object]) -> str:
            value = row["Real CPU total usage (%)"]
            return f"{value:.1f}%" if isinstance(value, float) else str(value)
        body = "".join(
            "<tr>"
            f"<td class='cpu-id'>{row['CPU ID']}</td>"
            f"<td class='usage'><strong>{row['FFmpeg avg usage (%)']:.1f}%</strong><span class='meter'><i style='width:{min(100.0, float(row['FFmpeg avg usage (%)'])):.1f}%'></i></span></td>"
            f"<td>{row['Encoder avg usage (%)']:.1f}%</td>"
            f"<td>{real_usage_text(row)}</td>"
            f"<td>{row['Decoder avg usage (%)']:.1f}%</td>"
            f"<td>{row['FFmpeg CPU time (s)']:.3f}</td>"
            f"<td>{row['Encoder CPU time (s)']:.3f}</td>"
            f"<td>{row['Decoder CPU time (s)']:.3f}</td>"
            "</tr>"
            for row in rows
        )
        return (
            "<div class='scroll core-table'><table><thead><tr>"
            "<th>Logical CPU ID</th><th>FFmpeg avg usage</th><th>Encoder avg usage</th>"
            "<th>Real CPU total usage</th><th>Decoder avg usage</th><th>FFmpeg CPU time (s)</th>"
            "<th>Encoder CPU time (s)</th><th>Decoder CPU time (s)</th>"
            f"</tr></thead><tbody>{body}</tbody></table></div>"
        )
    rdt_focus_session = config.get("RDT_FOCUS_SESSION", "").strip()
    rdt_scope = (
        f"single session {rdt_focus_session}; encoder and decoder groups only"
        if rdt_focus_session else
        f"all {config.get('STREAMS', '?')} sessions; one group per role"
    )
    basic = [
        {"Field": "Scenario", "Value": config.get("SCENARIO", "")},
        {"Field": "CPU placement", "Value": config.get("PLACEMENT", "")},
        {"Field": "Worker", "Value": config.get("NODE", config.get("LAB_DEFAULT_NODE", ""))},
        {"Field": "Resolution", "Value": config.get("RESOLUTION", "")},
        {"Field": "Preset", "Value": config.get("PRESET", "")},
        {"Field": "Encoder threads / slices", "Value": f"{config.get('ENC_THREADS', 'auto')} / {config.get('SLICES', 'auto')}"},
        {"Field": "Measurement window", "Value": f"{window.get('duration_s', '')} seconds"},
        {"Field": "Noisy neighbor", "Value": noise.get("profile", "disabled") if noise.get("enabled") else "disabled"},
        {"Field": "Noisy-neighbor scope", "Value": noise.get("scope", "n/a")},
        {"Field": "RDT monitoring", "Value": "enabled" if config.get("RDT_MONITOR") == "1" or config.get("RDT_CONTROL_PROFILE", "none") != "none" else "disabled"},
        {"Field": "RDT attribution scope", "Value": rdt_scope},
        {"Field": "RDT control profile", "Value": config.get("RDT_CONTROL_PROFILE", "none")},
        {"Field": "RDT state restored", "Value": rdt_stop.get("restored", "n/a")},
        {"Field": "stress-ng image", "Value": noise.get("image", "n/a")},
        {"Field": "stress-ng arguments", "Value": " ".join(noise.get("args", [])) or "n/a"},
        {"Field": "stress-ng cgroup CPU time", "Value": f"{noise_cpu_seconds:.3f} seconds" if noise_cpu_seconds is not None else "n/a"},
        {"Field": "Performance interpretation", "Value": interpretation},
    ]
    commands_html = "".join(
        f"<h3>{html.escape(row['pod'])} ({html.escape(row['role'])})</h3><pre>{html.escape(row['command'])}</pre>"
        for row in command_lines
    ) or "<p>No command-line evidence.</p>"
    document = f"""<!doctype html><html><head><meta charset='utf-8'><title>MXL performance report</title>
<style>:root{{--navy:#123b58;--blue:#1f6f9f;--line:#d7e1e8;--soft:#f4f8fb;--text:#18212b}}*{{box-sizing:border-box}}body{{font:14px system-ui;margin:0;background:#eef3f6;color:var(--text)}}main{{max-width:1500px;margin:auto;background:white;min-height:100vh;padding:2rem 3rem}}h1{{color:var(--navy);margin:0}}h2{{color:var(--navy);margin-top:2.4rem;border-bottom:2px solid #dce8ef;padding-bottom:.35rem}}.subtitle{{color:#607482;margin:.35rem 0 1.6rem}}.cards{{display:grid;grid-template-columns:repeat(auto-fit,minmax(180px,1fr));gap:1rem;margin:1rem 0 1.8rem}}.card{{background:var(--soft);border:1px solid var(--line);border-left:5px solid var(--blue);border-radius:8px;padding:1rem;box-shadow:0 2px 5px #17364d12}}.card span{{display:block;color:#607482;font-size:.82rem;text-transform:uppercase;letter-spacing:.04em}}.card strong{{display:block;color:var(--navy);font-size:1.55rem;margin-top:.25rem}}.card.pass{{border-left-color:#22863a;background:#f0faf3}}.card.pass strong{{color:#16712d}}.card.fail{{border-left-color:#cb2431;background:#fff3f4}}.card.fail strong{{color:#b31d28}}table{{border-collapse:separate;border-spacing:0;width:100%;border:1px solid var(--line);border-radius:6px;overflow:hidden}}th{{background:var(--navy);color:white;position:sticky;top:0}}th,td{{border-bottom:1px solid var(--line);padding:.55rem .65rem;text-align:left;vertical-align:top}}tr:last-child td{{border-bottom:0}}tr:nth-child(even){{background:var(--soft)}}.scroll{{overflow:auto;max-height:45rem}}pre{{background:#13232e;color:#e7f2f8;padding:1rem;border-radius:6px;overflow:auto}}code{{background:#eef3f6;padding:.1rem .3rem}}.method-note{{background:#eef7fb;border:1px solid #c8e2ed;border-left:5px solid var(--blue);border-radius:7px;padding:.8rem 1rem;color:#3f5968;line-height:1.5;margin:.8rem 0}}.method-note strong{{color:var(--navy)}}.core-table td:not(:first-child){{font-variant-numeric:tabular-nums}}.core-table .cpu-id{{font-weight:800;color:var(--navy)}}.usage{{min-width:190px}}.usage strong{{display:inline-block;width:52px}}.meter{{display:inline-block;width:105px;height:8px;background:#dce6eb;border-radius:99px;overflow:hidden;vertical-align:middle}}.meter i{{display:block;height:100%;background:linear-gradient(90deg,#25a5c5,#176b91);border-radius:inherit}}</style></head><body><main>
<h1>FFmpeg + MXL performance report</h1><p class='subtitle'>{html.escape(str(result.name))}</p>
<section class='metric-group'><h2>Result and throughput</h2><div class='cards'>{render_cards(throughput_cards)}</div></section>
<section class='metric-group'><h2>CPU</h2><div class='cards'>{render_cards(cpu_cards)}</div></section>
<section class='metric-group'><h2>Worker memory and bandwidth (host-wide)</h2><div class='cards'>{render_cards(memory_cards)}</div></section>
<section class='metric-group'><h2>Worker L3 cache (host-wide)</h2><div class='cards'>{render_cards(cache_cards)}</div></section>
<section class='metric-group'><h2>NUMA locality and cross-socket memory</h2><div class='cards'>{render_cards(numa_cards)}</div>
<div class='method-note'><strong>Meaning:</strong> Cross-socket UPI incoming bandwidth directly sums incoming traffic on all available UPI links and sockets once, without adding outgoing copies. Per-stream value divides same direct host-wide bandwidth by benchmark stream count for normalized comparison. Valid comparison requires same worker with no unrelated traffic. It is traffic intensity, not percentage of remote memory accesses or per-process attribution. Core-derived PCM local/remote ratio is intentionally excluded from headline results because it was not workload-sensitive under CPU Manager. Linux NUMA counters are exact-window whole-worker page-allocation deltas; they are counts, not bandwidth.</div>
{table(socket_numa_bandwidth)}
<h3>Exact-window NUMA allocation counters</h3>{table(numa_counters)}</section>
<h2>Basic test information</h2>{table(basic)}
<h2>Worker platform specification</h2>
<div class='method-note'><strong>Meaning:</strong> Core and memory topology probed from the worker itself. Theoretical DRAM peak is populated channels x transfer rate x 8 bytes; it is a hardware ceiling from configuration, never a measured or achievable rate. Real achievable bandwidth is typically well below it. Transfer rate comes from DMI when readable, otherwise from the declared <code>&lt;NODE&gt;_MEM_TRANSFER_MT_S</code> value in <code>config/nodes.env</code>; without either, peak stays unavailable rather than guessed.</div>
{table(platform_rows)}
<h2>RDT workload/noise attribution</h2>
<div class='method-note'><strong>Scope:</strong> {html.escape(rdt_scope)}. Each row is one role group on one L3 domain, not one FFmpeg process; add both domain rows only for that role-group total. LLC occupancy is footprint, not hit ratio. MBM is bandwidth attributed to monitored task groups. PCM remains host-wide source for LLC hit/miss, memory-controller, and UPI totals. Empty table means RDT was disabled.</div>
{table(rdt_summary)}
<h2>Noisy-neighbor evidence</h2><pre>{html.escape(json.dumps({"configuration": noise, "before": noise_before, "after": noise_after, "log": noise_log}, indent=2))}</pre>
<h2>Exact FFmpeg command lines</h2>{commands_html}
<h2>Exact process CPU affinity</h2>{table(placement)}
<h2>Per-core FFmpeg-only CPU usage</h2>
<div class='method-note'><strong>Meaning:</strong> FFmpeg and role columns estimate average process usage on each logical CPU during {measurement_duration:g}-second measurement window. CPU time comes from Linux <code>/proc/&lt;pid&gt;/task/&lt;tid&gt;/stat</code>; migrations between {config.get('LAB_SAMPLE_STEP', 5)}-second samples may shift attribution and make these estimates exceed 100%. <strong>Real CPU total usage</strong> is exact-window logical-CPU busy time from start/end <code>/proc/stat</code> counters. It is always 0–100% and includes FFmpeg plus Kubernetes, kernel, exporters, and all other host work. Comparing real total against encoder estimate shows how fully each logical CPU was occupied, but Linux standard accounting cannot isolate exact per-CPU FFmpeg share. With SMT on, a physical core appears here as its two sibling CPUs.</div>
{core_usage_table(core_usage)}
<h2>Per-process per-core CPU detail</h2>{table([{k: v for k, v in row.items() if k != "attribution"} for row in usage])}
<h2>All metrics</h2>{table(metrics)}
<h2>Host configuration</h2><pre>{html.escape(json.dumps(host, indent=2))}</pre>
<h2>Effective test configuration</h2><pre>{html.escape(json.dumps(config, indent=2, sort_keys=True))}</pre>
</main></body></html>"""
    (result / "report.html").write_text(document)
