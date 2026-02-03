"""Excel report generation for MTL nightly test reports."""

from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# Color scheme constants
COLOR_HEADER = "366092"
COLOR_PASSED = "d4edda"
COLOR_FAILED = "f8d7da"
COLOR_SKIPPED = "fff3cd"
COLOR_SEPARATOR = "e6f2ff"
COLOR_WHITE = "FFFFFF"


def generate_excel_report(pytest_data, gtest_data, output_file, system_info_list=None, test_metadata=None):
    """Create Excel report with separate sheets for pytest and gtest."""
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Create sheets
    if pytest_data:
        _create_pytest_sheet(wb, pytest_data)
    if gtest_data:
        _create_gtest_sheet(wb, gtest_data)
    _create_summary_sheet(wb, pytest_data, gtest_data, system_info_list, test_metadata)
    
    wb.save(output_file)
    print(f"Excel report saved to: {output_file}")


def _create_pytest_sheet(wb, pytest_data):
    """Create pytest results sheet with summary and detailed test cases."""
    ws = wb.create_sheet("Pytest Results")
    
    # Section 1: Summary Table
    _add_section_title(ws, 'PYTEST SUMMARY - BY CATEGORY', 'A1:I1')
    ws.append([])  # Blank row
    
    # Summary table
    summary_headers = ['NIC', 'Category', 'Passed', 'Failed', 'Skipped', 'Error', 'XPassed', 'XFailed', 'Total']
    ws.append(summary_headers)
    _style_header_row(ws, ws.max_row, 9)
    
    # Add summary data
    sorted_pytest = sorted(pytest_data, key=lambda x: (x.get('nic', ''), x.get('category', '')))
    for data in sorted_pytest:
        ws.append([
            data.get('nic', ''),
            data.get('category', ''),
            data.get('passed', 0),
            data.get('failed', 0),
            data.get('skipped', 0),
            data.get('error', 0),
            data.get('xpassed', 0),
            data.get('xfailed', 0),
            data.get('total', 0)
        ])
    
    # Section 2: Detailed Test Cases
    ws.append([])
    ws.append([])
    total_test_cases = sum(len(data.get('test_cases', [])) for data in pytest_data)
    print(f"Adding {total_test_cases} pytest detailed test cases to Excel")
    
    detail_title_row = ws.max_row + 1
    _add_section_title(ws, 'DETAILED TEST CASES', f'A{detail_title_row}:F{detail_title_row}')
    ws.append([])  # Blank row
    
    # Detailed table headers
    detail_headers = ['NIC', 'Category', 'Test Name', 'Result', 'Duration', 'Platform']
    ws.append(detail_headers)
    _style_header_row(ws, ws.max_row, 6)
    
    # Add all detailed test cases grouped by category
    for data in sorted_pytest:
        if 'test_cases' in data and data['test_cases']:
            _add_category_separator(ws, f"Category: {data.get('nic', '')} - {data.get('category', '')}", 6)
            
            sorted_cases = sorted(data['test_cases'], key=lambda x: x['test_name'])
            for tc in sorted_cases:
                ws.append([
                    tc['nic'],
                    tc['category'],
                    tc['test_name'],
                    tc['result'],
                    tc['duration'],
                    tc['platform']
                ])
                _color_code_result(ws, ws.max_row, 4, tc['result'])
    
    _auto_adjust_columns(ws)


def _create_gtest_sheet(wb, gtest_data):
    """Create gtest results sheet with summary and detailed test cases."""
    ws = wb.create_sheet("GTest Results")
    
    # Section 1: Summary Table
    _add_section_title(ws, 'GTEST SUMMARY - BY CATEGORY', 'A1:F1')
    ws.append([])  # Blank row
    
    # Summary table
    summary_headers = ['NIC', 'Test Category', 'Passed', 'Failed', 'Skipped', 'Total']
    ws.append(summary_headers)
    _style_header_row(ws, ws.max_row, 6)
    
    # Add summary data
    sorted_gtest = sorted(gtest_data, key=lambda x: (x.get('nic', ''), x.get('category', '')))
    for data in sorted_gtest:
        ws.append([
            data.get('nic', ''),
            data.get('category', ''),
            data.get('passed', 0),
            data.get('failed', 0),
            data.get('skipped', 0),
            data.get('total', 0)
        ])
    
    # Section 2: Detailed Test Cases
    ws.append([])
    ws.append([])
    
    detail_title_row = ws.max_row + 1
    _add_section_title(ws, 'DETAILED TEST CASES', f'A{detail_title_row}:F{detail_title_row}')
    ws.append([])  # Blank row
    
    # Detailed table headers
    detail_headers = ['NIC', 'Test Category', 'Test Name', 'Result', 'Duration', 'Platform']
    ws.append(detail_headers)
    _style_header_row(ws, ws.max_row, 6)
    
    # Add all detailed test cases grouped by category
    for data in sorted_gtest:
        if 'test_cases' in data and data['test_cases']:
            _add_category_separator(ws, f"Category: {data.get('nic', '')} - {data.get('category', '')}", 6)
            
            sorted_cases = sorted(data['test_cases'], key=lambda x: x['test_name'])
            for tc in sorted_cases:
                ws.append([
                    tc['nic'],
                    tc['category'],
                    tc['test_name'],
                    tc['result'],
                    tc['duration'],
                    tc['platform']
                ])
                _color_code_result(ws, ws.max_row, 4, tc['result'])
    
    _auto_adjust_columns(ws)


def _create_summary_sheet(wb, pytest_data, gtest_data, system_info_list, test_metadata):
    """Create summary sheet with overall statistics and test run metadata."""
    ws = wb.create_sheet("Summary", 0)
    
    summary_data = []
    summary_data.append(['MTL Nightly Test Report Summary'])
    summary_data.append(['Generated:', datetime.now().strftime('%Y-%m-%d %H:%M:%S UTC')])
    summary_data.append([])
    
    # Test run metadata
    if test_metadata:
        summary_data.extend(_build_metadata_section(test_metadata))
    
    # Calculate statistics
    stats = _calculate_statistics(pytest_data, gtest_data)
    
    # Add pytest summary
    if pytest_data:
        summary_data.extend(_build_test_summary('Pytest Summary', stats['pytest']))
    
    # Add gtest summary
    if gtest_data:
        summary_data.extend(_build_test_summary('GTest Summary', stats['gtest']))
    
    # Add overall summary
    if pytest_data or gtest_data:
        summary_data.extend(_build_test_summary('Overall Summary (Pytest + GTest)', stats['combined']))
    
    # Write all summary data
    for row in summary_data:
        ws.append(row)
    
    # Add system information
    if system_info_list:
        _add_system_info_section(ws, system_info_list)
    
    # Style summary sheet
    ws['A1'].font = Font(bold=True, size=14)
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    if system_info_list:
        for col_idx in range(3, 10):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 15


def _calculate_statistics(pytest_data, gtest_data):
    """Calculate test statistics from data."""
    pytest_total = pytest_passed = pytest_failed = pytest_skipped = 0
    gtest_total = gtest_passed = gtest_failed = gtest_skipped = 0
    
    if pytest_data:
        df = pd.DataFrame([{k: v for k, v in d.items() if k != 'test_cases'} for d in pytest_data])
        pytest_total = df['total'].sum()
        pytest_passed = df['passed'].sum()
        pytest_failed = df['failed'].sum()
        pytest_skipped = df['skipped'].sum()
    
    if gtest_data:
        df = pd.DataFrame([{k: v for k, v in d.items() if k != 'test_cases'} for d in gtest_data])
        gtest_total = df['total'].sum()
        gtest_passed = df['passed'].sum()
        gtest_failed = df['failed'].sum()
        gtest_skipped = df['skipped'].sum()
    
    pytest_pass_rate = (pytest_passed / (pytest_passed + pytest_failed) * 100) if (pytest_passed + pytest_failed) > 0 else 0
    gtest_pass_rate = (gtest_passed / (gtest_passed + gtest_failed) * 100) if (gtest_passed + gtest_failed) > 0 else 0
    
    combined_total = pytest_total + gtest_total
    combined_passed = pytest_passed + gtest_passed
    combined_failed = pytest_failed + gtest_failed
    combined_skipped = pytest_skipped + gtest_skipped
    combined_pass_rate = (combined_passed / (combined_passed + combined_failed) * 100) if (combined_passed + combined_failed) > 0 else 0
    
    return {
        'pytest': {
            'total': pytest_total,
            'passed': pytest_passed,
            'failed': pytest_failed,
            'skipped': pytest_skipped,
            'pass_rate': pytest_pass_rate
        },
        'gtest': {
            'total': gtest_total,
            'passed': gtest_passed,
            'failed': gtest_failed,
            'skipped': gtest_skipped,
            'pass_rate': gtest_pass_rate
        },
        'combined': {
            'total': combined_total,
            'passed': combined_passed,
            'failed': combined_failed,
            'skipped': combined_skipped,
            'pass_rate': combined_pass_rate
        }
    }


def _build_metadata_section(test_metadata):
    """Build test run metadata section."""
    data = [['Test Run Information'], []]
    
    if test_metadata.get('pytest_run_number') and test_metadata.get('pytest_branch'):
        run_date = test_metadata['pytest_run_date'].split('T')[0] if test_metadata.get('pytest_run_date') else 'N/A'
        run_time = test_metadata['pytest_run_date'].split('T')[1].split('Z')[0] if test_metadata.get('pytest_run_date') and 'T' in test_metadata['pytest_run_date'] else 'N/A'
        data.append(['Pytest Run:', f"#{test_metadata['pytest_run_number']} (branch: {test_metadata['pytest_branch']})"])
        data.append(['Pytest Date:', f'{run_date} {run_time}'])
        if test_metadata.get('pytest_run_url'):
            data.append(['Pytest URL:', test_metadata['pytest_run_url']])
    
    if test_metadata.get('gtest_run_number') and test_metadata.get('gtest_branch'):
        run_date = test_metadata['gtest_run_date'].split('T')[0] if test_metadata.get('gtest_run_date') else 'N/A'
        run_time = test_metadata['gtest_run_date'].split('T')[1].split('Z')[0] if test_metadata.get('gtest_run_date') and 'T' in test_metadata['gtest_run_date'] else 'N/A'
        data.append(['GTest Run:', f"#{test_metadata['gtest_run_number']} (branch: {test_metadata['gtest_branch']})"])
        data.append(['GTest Date:', f'{run_date} {run_time}'])
        if test_metadata.get('gtest_run_url'):
            data.append(['GTest URL:', test_metadata['gtest_run_url']])
    
    data.extend([[], []])
    return data


def _build_test_summary(title, stats):
    """Build test summary section."""
    return [
        [title],
        ['Total Tests:', stats['total']],
        ['Total Passed:', stats['passed']],
        ['Total Failed:', stats['failed']],
        ['Total Skipped:', stats['skipped']],
        ['Pass Rate:', f"{stats['pass_rate']:.2f}%"],
        []
    ]


def _add_system_info_section(ws, system_info_list):
    """Add system information section to worksheet."""
    ws.append([])
    ws.append([])
    ws.append(['Test Environment Information'])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)
    ws.append([])
    
    # Headers
    sys_headers = ['Hostname', 'Platform', 'CPU', 'Cores', 'RAM', 'HugePages', 'OS', 'Kernel', 'NICs']
    ws.append(sys_headers)
    _style_header_row(ws, ws.max_row, len(sys_headers))
    
    # Data rows
    for sys_info in system_info_list:
        ws.append([
            sys_info.get('hostname', 'unknown'),
            sys_info.get('platform', 'unknown'),
            sys_info.get('cpu', 'unknown'),
            sys_info.get('cpu_cores', 'unknown'),
            sys_info.get('ram', 'unknown'),
            sys_info.get('hugepages', 'unknown'),
            sys_info.get('os', 'unknown'),
            sys_info.get('kernel', 'unknown'),
            sys_info.get('nics', 'unknown')
        ])


def _add_section_title(ws, title, merge_range):
    """Add styled section title."""
    ws.append([title])
    cell = ws.cell(row=ws.max_row, column=1)
    cell.font = Font(bold=True, size=14, color=COLOR_WHITE)
    cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
    try:
        ws.merge_cells(merge_range)
    except:
        pass  # Ignore merge errors


def _style_header_row(ws, row_num, col_count):
    """Apply header styling to a row."""
    header_fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
    header_font = Font(bold=True, color=COLOR_WHITE)
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')


def _add_category_separator(ws, text, col_count):
    """Add category separator row."""
    sep_row = ws.max_row + 1
    row_data = [text] + [''] * (col_count - 1)
    ws.append(row_data)
    for col in range(1, col_count + 1):
        cell = ws.cell(row=sep_row, column=col)
        cell.font = Font(bold=True, italic=True)
        cell.fill = PatternFill(start_color=COLOR_SEPARATOR, end_color=COLOR_SEPARATOR, fill_type="solid")


def _color_code_result(ws, row_num, col_num, result):
    """Apply color coding to result cell."""
    cell = ws.cell(row=row_num, column=col_num)
    result_upper = result.upper()
    
    if 'PASS' in result_upper and 'FAIL' not in result_upper:
        cell.fill = PatternFill(start_color=COLOR_PASSED, end_color=COLOR_PASSED, fill_type="solid")
    elif 'FAIL' in result_upper or 'ERROR' in result_upper:
        cell.fill = PatternFill(start_color=COLOR_FAILED, end_color=COLOR_FAILED, fill_type="solid")
    elif 'SKIP' in result_upper:
        cell.fill = PatternFill(start_color=COLOR_SKIPPED, end_color=COLOR_SKIPPED, fill_type="solid")


def _auto_adjust_columns(ws):
    """Auto-adjust column widths based on content."""
    for col_idx, column in enumerate(ws.columns, start=1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 80)
